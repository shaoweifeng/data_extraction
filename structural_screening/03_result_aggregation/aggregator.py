import json
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
import rispy

all_df = None
_xml_cache = {}


def _itext(elem):
    if elem is None:
        return ""
    return "".join(elem.itertext()).strip()


def _load_xml_fields(xml_path):
    if not xml_path:
        return {}
    cached = _xml_cache.get(xml_path)
    if cached is not None:
        return cached
    try:
        import xml.etree.ElementTree as ET
        root = ET.parse(xml_path).getroot()
        ref = None
        if root.tag == "Reference":
            ref = root
        else:
            ref = root.find(".//Reference")
        if ref is None:
            _xml_cache[xml_path] = {}
            return {}

        authors = ", ".join([_itext(a) for a in ref.findall("./Authors/Author") if _itext(a)])
        fields = {
            "ReferenceType": _itext(ref.find("ReferenceType")),
            "Volume": _itext(ref.find("Volume")),
            "Issue": _itext(ref.find("Issue")),
            "Page": _itext(ref.find("Page")),
            "Date": _itext(ref.find("Date")),
            "Doi": _itext(ref.find("DOI")),
            "PMCID": _itext(ref.find("PMCID")),
            "Abstract": _itext(ref.find("Abstract")),
            "Address": _itext(ref.find("Address")),
            "Title": _itext(ref.find("Title")),
            "Author": authors,
            "Journal": _itext(ref.find("Journal")),
            "Year": _itext(ref.find("Year")),
            "URL": _itext(ref.find("URL")),
        }
        _xml_cache[xml_path] = fields
        return fields
    except Exception:
        _xml_cache[xml_path] = {}
        return {}


def simple_json_to_excel(json_file_path, cnt):
    """
    简化版：不需要Excel模板文件，直接使用已知的表头顺序
    """
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"读取JSON失败，跳过: {json_file_path} ({e})")
        return False

    # 定义基础表头顺序
    base_headers = [
        "id",
        "include_or_not", "exclusion_reason_id", "exclusion_reason",
        "ReferenceType",
        "Title", "Author", "Year", "Journal",
        "Volume", "Issue", "Page", "Date", "Doi", "PMCID", "Abstract", "URL", "Address",
        "source_xml"
    ]
    
    # 收集所有动态出现的字段
    dynamic_headers = set()

    # 字段映射（JSON中的字段名 -> 表头字段名）
    # screener.py 生成的结构已经是扁平化的，或者在 prompt1_result 中
    # prompt1_result: [{"title": ..., "authors": [...], "exclusion_reason": ...}]
    field_mapping = {
        'first_author': 'Author', # 兼容旧
        'authors': 'Author',
        'public_year': 'Year',     # 兼容旧
        'year': 'Year',
        'file_path': 'source_xml',  # 兼容旧
        'title': 'Title',
        'journal': 'Journal',
        'url': 'URL',
        'doi': 'Doi',
        'extracted_abstract': 'Abstract',
        'number_exclusion_reason': 'exclusion_reason_id'
    }

    # 处理prompt1_result（基本信息）
    # 兼容两种格式：
    # 1. 嵌套格式：data['prompt1_result']
    # 2. 扁平格式：data 本身就是结果字典
    
    prompt1_data = data.get('prompt1_result', [])
    
    # 如果 prompt1_result 为空，但 data 本身包含 'include_or_not' 等关键字段，则视为扁平格式
    if not prompt1_data and ('include_or_not' in data or 'exclusion_reason' in data):
        prompt1_data = [data]

    # 构建最终数据
    all_rows = []

    # 如果有prompt1_data，将其作为基础信息
    if prompt1_data:
        # prompt1_data 可能是 list 或 dict
        base_info_list = prompt1_data if isinstance(prompt1_data, list) else [prompt1_data]
        
        for base_info in base_info_list:
            if not isinstance(base_info, dict): continue
            
            # 应用字段映射
            mapped_base_info = {}
            # 先复制所有字段，确保 URL 等未显式映射的字段也被包含
            mapped_base_info.update(base_info)
            
            for key, value in base_info.items():
                mapped_key = field_mapping.get(key, key)
                # 处理列表转字符串 (如 authors)
                if isinstance(value, list):
                    # 只有当 value 是 list 时才 join，避免报错
                    try:
                        value = ", ".join(str(v) for v in value)
                    except:
                        pass
                mapped_base_info[mapped_key] = value
                
                # 记录动态字段 (排除已知的基础字段)
                if mapped_key not in base_headers:
                    dynamic_headers.add(mapped_key)
            
            # 确保包含 source_xml，如果没有则尝试从外层获取
            if 'source_xml' not in mapped_base_info and 'pdf_file' in data:
                 mapped_base_info['source_xml'] = data['pdf_file']
            
            # 确保包含 url，如果内层没有，尝试从外层获取 (针对扁平结构可能不需要，但为了保险)
            if 'URL' not in mapped_base_info and 'url' in data:
                 mapped_base_info['URL'] = data['url']

            source_xml = mapped_base_info.get('source_xml') or data.get('source_xml') or ""
            results_dir = os.path.dirname(os.path.dirname(json_file_path))
            screening_ai_dir = os.path.dirname(results_dir)
            datasets_dir = os.path.join(screening_ai_dir, "datasets")
            xml_path = os.path.join(datasets_dir, os.path.basename(source_xml)) if source_xml else ""
            xml_fields = _load_xml_fields(xml_path) if xml_path and os.path.exists(xml_path) else {}

            if 'ReferenceType' not in mapped_base_info:
                mapped_base_info['ReferenceType'] = xml_fields.get('ReferenceType')
            if 'Title' not in mapped_base_info:
                mapped_base_info['Title'] = xml_fields.get('Title')
            if 'Author' not in mapped_base_info:
                mapped_base_info['Author'] = xml_fields.get('Author')
            if 'Year' not in mapped_base_info:
                mapped_base_info['Year'] = xml_fields.get('Year')
            if 'Journal' not in mapped_base_info:
                mapped_base_info['Journal'] = xml_fields.get('Journal')
            if 'Volume' not in mapped_base_info:
                mapped_base_info['Volume'] = xml_fields.get('Volume')
            if 'Issue' not in mapped_base_info:
                mapped_base_info['Issue'] = xml_fields.get('Issue')
            if 'Page' not in mapped_base_info:
                mapped_base_info['Page'] = xml_fields.get('Page')
            if 'Date' not in mapped_base_info:
                mapped_base_info['Date'] = xml_fields.get('Date')
            if 'Doi' not in mapped_base_info:
                mapped_base_info['Doi'] = xml_fields.get('Doi') or data.get('doi') or mapped_base_info.get('doi')
            if 'PMCID' not in mapped_base_info:
                mapped_base_info['PMCID'] = xml_fields.get('PMCID')
            if 'Abstract' not in mapped_base_info:
                mapped_base_info['Abstract'] = xml_fields.get('Abstract') or data.get('extracted_abstract') or mapped_base_info.get('extracted_abstract')
            if 'URL' not in mapped_base_info:
                mapped_base_info['URL'] = xml_fields.get('URL') or data.get('url') or mapped_base_info.get('url')
            if 'Address' not in mapped_base_info:
                mapped_base_info['Address'] = xml_fields.get('Address')

            all_rows.append(mapped_base_info)

    # 创建DataFrame
    if all_rows:
        df = pd.DataFrame(all_rows)
        
        # 合并表头：基础字段 + 排序后的动态字段
        # final_headers = base_headers + sorted(list(dynamic_headers))
        # 仅使用基础表头，忽略动态列
        final_headers = base_headers

        # 确保所有表头列都存在
        for header in final_headers:
            if header not in df.columns:
                df[header] = None

        # 按照表头顺序重新排列列
        # df['id'] = cnt # 如果需要自增ID
        if 'id' in df.columns:
             df['id'] = cnt # 覆盖 id
        else:
             df['id'] = cnt
        
        # 只保留 final_headers 中的列，并按顺序排列
        # 使用 reindex 可以安全地处理缺失列（自动填 NaN）
        df = df.reindex(columns=final_headers)

        global all_df
        if all_df is None:
            all_df = df.copy()
        else:
            # 合并时处理列不一致的情况
            all_df = pd.concat([all_df, df], axis=0, ignore_index=True)

        return True
    else:
        print("没有找到可用的数据，filepath：", json_file_path)
        return False

def convert_to_ris(json_dir, output_ris_path):
    """
    将筛选结果转换为 RIS 格式 (EndNote)
    """
    files = traverse_directory(json_dir)
    ris_entries = []
    
    for file_path in files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            continue
            
        # 提取信息 (假设 prompt1_result 包含元数据)
        prompt1_data = data.get('prompt1_result', [])
        if not prompt1_data:
            continue
            
        base_info = prompt1_data[0]
        
        # 映射到 RIS 字段
        entry = {
            'type_of_reference': 'JOUR',
            'title': base_info.get('title'),
            'authors': [base_info.get('first_author')] if base_info.get('first_author') else [],
            'year': base_info.get('public_year'),
            # 'secondary_title': base_info.get('journal'), # 如果有期刊名
            # 'abstract': base_info.get('abstract'), # 如果有摘要
            'custom1': data.get('pdf_file') # 自定义字段存文件名
        }
        
        # 清理空值
        entry = {k: v for k, v in entry.items() if v}
        ris_entries.append(entry)
        
    if ris_entries:
        with open(output_ris_path, 'w', encoding='utf-8') as f:
            rispy.dump(ris_entries, f)
        print(f"RIS 文件已生成: {output_ris_path} (共 {len(ris_entries)} 条)")
        return True
    else:
        print("没有可生成的 RIS 条目")
        return False

def traverse_directory(path):
    """
    【修复】每个结果目录只读取最新的一个 JSON 文件，避免重复计算
    """
    from collections import defaultdict
    
    # 按目录分组收集 JSON 文件
    dir_files = defaultdict(list)
    for root, dirs, files in os.walk(path):
        for file in files:
            if re.search(r'\.json$', file, re.IGNORECASE):
                file_path = os.path.join(root, file)
                dir_files[root].append(file_path)
    
    # 每个目录只取最新的一个文件（按修改时间排序）
    file_pathes = []
    for dir_path, files in dir_files.items():
        if files:
            # 按修改时间降序排序，取最新的
            files_with_mtime = [(f, os.path.getmtime(f)) for f in files]
            files_with_mtime.sort(key=lambda x: x[1], reverse=True)
            file_pathes.append(files_with_mtime[0][0])
    
    return file_pathes


def merge_same_cells_advanced(excel_path, output_path):
    """
    智能合并方法：先按前4列分组，再合并组内完全相同的列
    """
    # 读取数据
    df = pd.read_excel(excel_path, sheet_name='Sheet1')

    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb.active

    # 取消所有已有的合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    total_rows = len(df) + 1  # +1 因为包含标题行
    total_cols = len(df.columns)

    print(f"表格总行数: {total_rows}, 总列数: {total_cols}")

    # 按行分组：前4列相同的行分为一组
    groups = []
    current_group = []

    for row_idx in range(2, total_rows + 1):  # 从第2行开始（数据行）
        if row_idx == 2:
            current_group = [row_idx]
            continue

        # 检查前4列是否相同
        same = True
        for col in range(1, 5):  # 前4列
            current_val = ws.cell(row=row_idx, column=col).value
            prev_val = ws.cell(row=row_idx - 1, column=col).value

            # 处理NaN值
            if current_val is None or (isinstance(current_val, float) and pd.isna(current_val)):
                current_val = None
            if prev_val is None or (isinstance(prev_val, float) and pd.isna(prev_val)):
                prev_val = None

            if current_val != prev_val:
                same = False
                break

        if same:
            current_group.append(row_idx)
        else:
            if len(current_group) > 1:
                groups.append(current_group)
            current_group = [row_idx]

    # 添加最后一组
    if len(current_group) > 1:
        groups.append(current_group)

    print(f"找到 {len(groups)} 个需要合并的组")

    # 对每个组内的每一列检查是否需要合并
    merge_info = []

    for i, group in enumerate(groups):
        start_row = group[0]
        end_row = group[-1]

        # print(f"处理第 {i + 1} 组: 行 {start_row} 到 {end_row}")

        for col_idx in range(1, total_cols + 1):
            # 检查该列在组内是否所有值都相同
            values_in_group = []
            for row in group:
                value = ws.cell(row=row, column=col_idx).value
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    value = None
                values_in_group.append(value)

            # 如果组内所有值都相同，则合并
            if len(set(values_in_group)) == 1 and len(group) > 1:
                col_letter = get_column_letter(col_idx)  # 使用正确的列字母转换函数
                merge_info.append((col_letter, start_row, end_row))
                # print(f"  列 {col_letter} 需要合并 (值: {values_in_group[0]})")

    # 执行合并操作（从后往前合并，避免索引问题）
    for col_letter, start_row, end_row in reversed(merge_info):
        try:
            merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            # print(f"合并范围: {merge_range}")
            ws.merge_cells(merge_range)

            # 设置居中对齐
            merged_cell = ws[f"{col_letter}{start_row}"]
            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except Exception as e:
            print(f"合并 {merge_range} 时出错: {e}")

    # 保存文件
    wb.save(output_path)
    print(f"合并完成，文件已保存到: {output_path}")
    print(f"共合并了 {len(merge_info)} 个单元格区域")

    return len(merge_info)


# 使用示例
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Aggregate screening results to Excel")
    parser.add_argument('--input_dir', type=str, help='Input directory containing JSON files', required=False)
    parser.add_argument('--output_dir', type=str, help='Output directory for Excel files', required=False)
    args = parser.parse_args()

    # 兼容旧逻辑：如果没有传参，尝试使用默认路径（但不建议，应强制传参）
    # 为了保证项目隔离，我们不应该 chdir，而是直接使用绝对路径
    
    # os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) # REMOVED
    
    input_dir = args.input_dir
    output_dir = args.output_dir
    
    if not input_dir:
        # Fallback (仅用于本地单独测试，生产环境 tasks.py 必须传参)
        # 假设在 structural_screening/03_result_aggregation 运行
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        input_dir = os.path.join(base_dir, "02_screening_ai", "results")
        print(f"Warning: No input_dir provided, using default: {input_dir}")

    if not output_dir:
         output_dir = os.path.join(os.path.dirname(input_dir), "results") # 默认输出到 input_dir 同级的 results
         print(f"Warning: No output_dir provided, using default: {output_dir}")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_file = os.path.join(output_dir, "AI初筛结果.xlsx")
    final_file = os.path.join(output_dir, "AI初筛结果_合并版.xlsx")
    
    if not os.path.exists(input_dir):
        print(f"Error: Input directory does not exist: {input_dir}")
        sys.exit(1)
    
    files = traverse_directory(input_dir)
    print(f"Found {len(files)} JSON files to aggregate")
    cnt, failed_cnt, succ_cnt = 0, 0, 0
    for file in files:
        assert isinstance(file, str)
        success = simple_json_to_excel(file, cnt+1)
        cnt = cnt + 1
        if not success:
            failed_cnt = failed_cnt + 1
            # print("file ", file, " not success, failed count = ", cnt)
        else:
            succ_cnt = succ_cnt + 1

    # 保存到Excel
    if all_df is None:
        print("警告：没有可导出的数据 (all_df is None)，生成空表格")
        all_df = pd.DataFrame(columns=["id", "title", "authors", "journal", "year", "exclusion_reason", "include_or_not", "url", "source_xml"])
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    all_df.to_excel(output_file, index=False, engine='openpyxl')
    print (f"成功写入{succ_cnt}篇，失败{failed_cnt}篇")
    print(f"数据已成功保存到 {output_file}")

    merge_same_cells_advanced(output_file, final_file)
