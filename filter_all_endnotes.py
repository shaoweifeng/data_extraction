#!/usr/bin/env python3
"""
遍历 endnotes/ 下所有 xml 文件（排除 embase_filtered.xml），
筛选出标题在 test.xlsx Title 列中出现的 record，
合并去重后写入 endnotes/all_filtered.xml，并打印统计信息。
"""

import os
import xml.etree.ElementTree as ET
import openpyxl
from collections import defaultdict

# ── 1. 读取 test.xlsx Title ───────────────────────────────────────────────────
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]
title_col_idx = headers.index("Title")

xlsx_titles = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[title_col_idx]
    if t:
        xlsx_titles.add(str(t).strip().lower())

print(f"test.xlsx 共有 {len(xlsx_titles)} 条标题\n")

# ── 2. 辅助函数 ───────────────────────────────────────────────────────────────
def get_record_title(record_el):
    titles_el = record_el.find("titles")
    if titles_el is None:
        return ""
    title_el = titles_el.find("title")
    if title_el is None:
        return ""
    return "".join(title_el.itertext()).strip()

# ── 3. 遍历所有 xml 文件 ──────────────────────────────────────────────────────
ENDNOTES_DIR = "endnotes"
EXCLUDE = {"embase_filtered.xml", "all_filtered.xml"}

xml_files = sorted([
    f for f in os.listdir(ENDNOTES_DIR)
    if f.endswith(".xml") and f not in EXCLUDE
])

print(f"找到 {len(xml_files)} 个 xml 文件: {xml_files}\n")

# title(lower) -> (record_el, source_file)
seen_titles = {}          # 去重用：已保留的 title -> source
kept_records = []         # 保留的 (record_el, source_file)

stats = []

for fname in xml_files:
    fpath = os.path.join(ENDNOTES_DIR, fname)
    try:
        tree = ET.parse(fpath)
    except ET.ParseError as e:
        print(f"  [WARN] {fname} 解析失败: {e}")
        stats.append((fname, 0, 0, 0, 0))
        continue

    root = tree.getroot()
    records_node = root.find("records")
    if records_node is None:
        # 有些文件 record 直接在根下
        all_recs = list(root.iter("record"))
    else:
        all_recs = list(records_node)

    total = len(all_recs)
    matched = 0
    dedup_skipped = 0

    for rec in all_recs:
        title = get_record_title(rec)
        title_lower = title.lower()
        if title_lower not in xlsx_titles:
            continue
        matched += 1
        if title_lower in seen_titles:
            dedup_skipped += 1
            continue
        seen_titles[title_lower] = fname
        kept_records.append((rec, fname))

    stats.append((fname, total, matched, dedup_skipped, matched - dedup_skipped))
    print(f"  {fname}: 共 {total} 条，匹配 {matched} 条，去重跳过 {dedup_skipped} 条，新增 {matched - dedup_skipped} 条")

# ── 4. 构建输出 XML ───────────────────────────────────────────────────────────
out_root = ET.Element("xml")
out_records = ET.SubElement(out_root, "records")
for rec, _ in kept_records:
    out_records.append(rec)

out_tree = ET.ElementTree(out_root)
ET.indent(out_tree, space="  ")
output_path = os.path.join(ENDNOTES_DIR, "all_filtered.xml")
out_tree.write(output_path, encoding="unicode", xml_declaration=True)

# ── 5. 汇总统计 ───────────────────────────────────────────────────────────────
total_all = sum(s[1] for s in stats)
total_matched = sum(s[2] for s in stats)
total_dedup = sum(s[3] for s in stats)
total_kept = len(kept_records)

print(f"""
══════════════════════════════════════
汇总统计
══════════════════════════════════════
处理文件数        : {len(xml_files)}
所有文件 record 数 : {total_all}
匹配 xlsx 标题数   : {total_matched}（含跨文件重复）
跨文件去重跳过     : {total_dedup}
最终保留 record 数 : {total_kept}

输出文件: {output_path}
══════════════════════════════════════

test.xlsx 中未在任何 xml 里找到的标题（{len(xlsx_titles) - total_kept} 条）:""")

found_titles = set(seen_titles.keys())
for t in sorted(xlsx_titles - found_titles):
    print(f"  - {t[:100]}")
