import json
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter

all_df = None


def simple_json_to_excel(json_file_path, cnt):
    """
    简化版：不需要Excel模板文件，直接使用已知的表头顺序
    """
    # 读取JSON文件
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 定义表头顺序（根据提供的Excel表头）
    header_order = [
 "title","1","2","3","4","5","6","7","8","9","10","11","12","13"
]


    # 字段映射（JSON中的字段名 -> 表头字段名）
    field_mapping = {
        'first_author': 'author',
        'public_year': 'year',
        'title': 'name',
        'CKMB_value': 'CK-MB_value'
    }

    # 处理prompt1_result（基本信息）
    prompt1_data = data.get('prompt1_result', [])

    # 处理prompt2_result（irAE记录）
    prompt2_data = data.get('prompt2_result', [])

    # 处理prompt3_result（细胞因子）
    prompt3_data = data.get('prompt3_result', {})  # 默认空字典，避免None
    # print(f"[DEBUG] prompt3_data 原始内容: {prompt3_data}")  # 调试打印

    # 标准化prompt3_data为字典格式
    if isinstance(prompt3_data, list):
        if len(prompt3_data) > 0:
            IL_data = prompt3_data[0] if isinstance(prompt3_data[0], dict) else {}
        else:
            IL_data = {}
    elif isinstance(prompt3_data, dict):
        IL_data = prompt3_data
    else:
        IL_data = {}

    # print(f"[DEBUG] 标准化后的 IL_data: {IL_data}")  # 调试打印

    # 构建最终数据
    all_rows = []

    # 如果有prompt1_data，将其作为基础信息
    if prompt1_data:
        base_info = prompt1_data[0]  # 取第一条作为基础信息

        # 应用字段映射
        mapped_base_info = {}
        for key, value in base_info.items():
            if value is None or value == 'null':
                continue
            mapped_key = field_mapping.get(key, key)
            mapped_base_info[mapped_key] = value

        # 如果有prompt2_data，为每条irAE记录创建一行
        if prompt2_data:
            for irAE_record in prompt2_data:
                # 应用字段映射到irAE记录
                mapped_irAE_record = {}
                for key, value in irAE_record.items():
                    if value is None or value == 'null':
                        continue
                    mapped_key = field_mapping.get(key, key)
                    mapped_irAE_record[mapped_key] = value

                # 合并基础信息和irAE记录
                row_data = mapped_base_info.copy()
                row_data.update(mapped_irAE_record)

                # 添加细胞因子检查信息
                if IL_data:  # 如果IL_data非空
                    row_data['cytokine_check'] = '有'
                    for key, value in IL_data.items():
                        if value is None or value == 'null':
                            continue
                        mapped_key = field_mapping.get(key, key)
                        row_data[mapped_key] = value
                else:
                    row_data['cytokine_check'] = '无'

                all_rows.append(row_data)
        else:
            # 如果没有irAE记录，只添加基础信息
            if IL_data:
                mapped_base_info['cytokine_check'] = '有'
                for key, value in IL_data.items():
                    if value is None or value == 'null':
                        continue
                    mapped_key = field_mapping.get(key, key)
                    mapped_base_info[mapped_key] = value
            else:
                mapped_base_info['cytokine_check'] = '无'

            all_rows.append(mapped_base_info)

    # 创建DataFrame
    if all_rows:
        df = pd.DataFrame(all_rows)

        # 确保所有表头列都存在
        for header in header_order:
            if header not in df.columns:
                df[header] = None

        # 按照表头顺序重新排列列
        df = df[header_order]
        # df['id'] = cnt    会在末尾
        df.insert(loc=0, column='id', value=cnt)    # 插到开头

        global all_df
        if all_df is None:
            all_df = df.copy()
        else:
            all_df = pd.concat([all_df, df], axis=0)

        return True
    else:
        print("没有找到可用的数据，filepath：", json_file_path)
        return False
def traverse_directory(path):
    file_pathes = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if re.search(r'\.json$', file, re.IGNORECASE):
                file_path = os.path.join(root, file)
                file_pathes.append(file_path)
    return file_pathes


def merge_same_cells_advanced(excel_path, output_path):
    """
    智能合并方法：先按前4列分组，再合并组内完全相同的列
    """
    # 读取数据
    df = pd.read_excel(excel_path, sheet_name='Sheet1')

    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb.active

    # 取消所有已有的合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    total_rows = len(df) + 1  # +1 因为包含标题行
    total_cols = len(df.columns)

    print(f"表格总行数: {total_rows}, 总列数: {total_cols}")

    # 按行分组：前4列相同的行分为一组
    groups = []
    current_group = []

    for row_idx in range(2, total_rows + 1):  # 从第2行开始（数据行）
        if row_idx == 2:
            current_group = [row_idx]
            continue

        # 检查前4列是否相同
        same = True
        for col in range(1, 5):  # 前4列
            current_val = ws.cell(row=row_idx, column=col).value
            prev_val = ws.cell(row=row_idx - 1, column=col).value

            # 处理NaN值
            if current_val is None or (isinstance(current_val, float) and pd.isna(current_val)):
                current_val = None
            if prev_val is None or (isinstance(prev_val, float) and pd.isna(prev_val)):
                prev_val = None

            if current_val != prev_val:
                same = False
                break

        if same:
            current_group.append(row_idx)
        else:
            if len(current_group) > 1:
                groups.append(current_group)
            current_group = [row_idx]

    # 添加最后一组
    if len(current_group) > 1:
        groups.append(current_group)

    print(f"找到 {len(groups)} 个需要合并的组")

    # 对每个组内的每一列检查是否需要合并
    merge_info = []

    for i, group in enumerate(groups):
        start_row = group[0]
        end_row = group[-1]

        # print(f"处理第 {i + 1} 组: 行 {start_row} 到 {end_row}")

        for col_idx in range(1, total_cols + 1):
            # 检查该列在组内是否所有值都相同
            values_in_group = []
            for row in group:
                value = ws.cell(row=row, column=col_idx).value
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    value = None
                values_in_group.append(value)

            # 如果组内所有值都相同，则合并
            if len(set(values_in_group)) == 1 and len(group) > 1:
                col_letter = get_column_letter(col_idx)  # 使用正确的列字母转换函数
                merge_info.append((col_letter, start_row, end_row))
                # print(f"  列 {col_letter} 需要合并 (值: {values_in_group[0]})")

    # 执行合并操作（从后往前合并，避免索引问题）
    for col_letter, start_row, end_row in reversed(merge_info):
        try:
            merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            # print(f"合并范围: {merge_range}")
            ws.merge_cells(merge_range)

            # 设置居中对齐
            merged_cell = ws[f"{col_letter}{start_row}"]
            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        except Exception as e:
            print(f"合并 {merge_range} 时出错: {e}")

    # 保存文件
    wb.save(output_path)
    print(f"合并完成，文件已保存到: {output_path}")
    print(f"共合并了 {len(merge_info)} 个单元格区域")

    return len(merge_info)


# 使用示例
if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # print(os.getcwd())
    output_file = "2/免疫文献提取结果.xlsx"
    final_file = "2/免疫文献提取结果_合并版.xlsx"

    # 使用简化版（不需要模板文件）
    files = traverse_directory('2/results')
    cnt, failed_cnt, succ_cnt = 0, 0, 0
    for file in files:
        assert isinstance(file, str)
        success = simple_json_to_excel(file, cnt+1)
        cnt = cnt + 1
        if not success:
            failed_cnt = failed_cnt + 1
            # print("file ", file, " not success, failed count = ", cnt)
        else:
            succ_cnt = succ_cnt + 1

    # 保存到Excel
    all_df.to_excel(output_file, index=False, engine='openpyxl')
    print (f"成功写入{succ_cnt}篇，失败{failed_cnt}篇")
    print(f"数据已成功保存到 {output_file}")

    merge_same_cells_advanced(output_file, final_file)