#!/usr/bin/env python3
"""
从 endnotes/embase.xml 中筛选出在 test.xlsx Title 列中出现的 record，
保留这些 record，删除其余，输出到 endnotes/embase_filtered.xml。
"""

import xml.etree.ElementTree as ET
import openpyxl

# ── 1. 读取 test.xlsx 中所有 Title（忽略大小写、去首尾空格）──────────────────
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]
title_col_idx = headers.index("Title")  # 0-based

xlsx_titles = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[title_col_idx]
    if t:
        xlsx_titles.add(str(t).strip().lower())

print(f"test.xlsx 共有 {len(xlsx_titles)} 条标题")

# ── 2. 解析 embase.xml ────────────────────────────────────────────────────────
ET.register_namespace("", "")  # 不改命名空间
tree = ET.parse("endnotes/embase.xml")
root = tree.getroot()         # <xml>
records_node = root.find("records")  # <records>

# ── 3. 提取 record 的标题（在 <titles><title><style>...</style></title></titles> 里）──
def get_record_title(record_el):
    titles_el = record_el.find("titles")
    if titles_el is None:
        return ""
    title_el = titles_el.find("title")
    if title_el is None:
        return ""
    # 拼接 title_el 下所有文本（含子元素文本）
    return "".join(title_el.itertext()).strip()

# ── 4. 筛选 ───────────────────────────────────────────────────────────────────
all_records = list(records_node)  # records 的直接子元素即 record 列表
kept = []
removed = []

for rec in all_records:
    title = get_record_title(rec)
    if title.lower() in xlsx_titles:
        kept.append(rec)
    else:
        removed.append(title)

print(f"embase.xml 共 {len(all_records)} 条 record")
print(f"保留：{len(kept)} 条，删除：{len(removed)} 条")

# ── 5. 重建 records 节点 ──────────────────────────────────────────────────────
for rec in all_records:
    records_node.remove(rec)

for rec in kept:
    records_node.append(rec)

# ── 6. 写出新文件 ─────────────────────────────────────────────────────────────
ET.indent(tree, space="  ")  # Python 3.9+，格式化缩进（可选）
output_path = "endnotes/embase_filtered.xml"
tree.write(output_path, encoding="unicode", xml_declaration=True)
print(f"已写出 {output_path}")
