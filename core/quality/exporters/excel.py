"""QA Excel 工作簿导出器，不依赖 HTTP request/response。"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

from core.models import QAReference


def export_qa_excel(project, quality_method, include_unconfirmed=False):
    refs = QAReference.objects.filter(project=project, quality_method=quality_method).prefetch_related('signal_items__confirmed_by', 'domain_results')

    wb = openpyxl.Workbook()

    # ── Sheet 1: 评价明细 ──────────────────────────────────
    ws = wb.active
    ws.title = '评价明细'
    headers = [
        '项目名称', '文献标题', '第一作者', '年份', '质量评价方法', '评价模式',
        '评价领域', '结果类型', '信号问题', '中文释义',
        '模型1判断', '模型1理由', '模型2判断', '模型2理由',
        '一致性状态', '系统推荐', 'AI判断', '判断理由',
        '人工最终判断', '是否人工修改', '确认人', '确认时间', '证据位置',
        '是否已确认',
    ]
    ws.append(headers)

    COLOR_MAP = {'low': 'C8F7C5', 'high': 'FFCDD2', 'unclear': 'FFF9C4', 'pending': 'F5F5F5'}

    for ref in refs:
        for item in ref.signal_items.all():
            if not include_unconfirmed and not item.is_confirmed:
                row_color = 'F5F5F5'
            else:
                row_color = None

            row = [
                project.name,
                ref.title,
                ref.first_author,
                ref.year,
                ref.quality_method,
                ref.get_eval_mode_display() if ref.eval_mode else '',
                item.domain,
                item.result_type,
                item.signal_question,
                item.signal_description,
                item.model1_judgment,
                item.model1_reason,
                item.model2_judgment,
                item.model2_reason,
                item.consistency,
                item.system_recommendation,
                item.ai_judgment,
                item.ai_reason,
                item.human_judgment,
                '是' if item.is_modified else '否',
                item.confirmed_by.username if item.confirmed_by else '',
                item.confirmed_at.strftime('%Y-%m-%d %H:%M') if item.confirmed_at else '',
                item.ai_evidence_page,
                '是' if item.is_confirmed else '否（未确认）',
            ]
            ws.append(row)
            if row_color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(fill_type='solid', fgColor=row_color)

    # ── Sheet 2: 汇总统计 ──────────────────────────────────
    ws2 = wb.create_sheet('汇总统计')
    wb.move_sheet(ws2, offset=-1)
    ws2.append(['文献标题', '第一作者', '年份', '患者选择_偏倚', '待评价试验_偏倚', '参考标准_偏倚', '流程与时间_偏倚',
                '患者选择_适用', '待评价试验_适用', '参考标准_适用', '整体审阅状态'])
    for ref in refs:
        dr_map = {dr.domain: dr for dr in ref.domain_results.all()}
        ps  = dr_map.get('patient_selection')
        it  = dr_map.get('index_test')
        rs  = dr_map.get('reference_standard')
        ft  = dr_map.get('flow_timing')
        ws2.append([
            ref.title, ref.first_author, ref.year,
            ps.bias_risk_result if ps else '',
            it.bias_risk_result if it else '',
            rs.bias_risk_result if rs else '',
            ft.bias_risk_result if ft else '',
            ps.applicability_result if ps else '',
            it.applicability_result if it else '',
            rs.applicability_result if rs else '',
            ref.review_status,
        ])

    # ── Sheet 3: 证据记录 ──────────────────────────────────
    ws3 = wb.create_sheet('证据记录')
    ws3.append(['文献标题', '信号问题', 'AI判断', '判断理由', '证据原文', '证据位置', '人工修改前', '人工最终判断', '是否修改'])
    for ref in refs:
        for item in ref.signal_items.filter(ai_evidence__gt=''):
            ws3.append([
                ref.title, item.signal_question,
                item.ai_judgment, item.ai_reason,
                item.ai_evidence, item.ai_evidence_page,
                item.original_ai_judgment, item.human_judgment,
                '是' if item.is_modified else '否',
            ])

    # ── Sheet 4: 多模型校验记录 ────────────────────────────
    ws4 = wb.create_sheet('多模型校验记录')
    ws4.append(['文献标题', '信号问题', '模型1 ID', '模型1判断', '模型1理由', '模型2 ID', '模型2判断', '模型2理由', '一致性', '系统推荐', '人工最终判断'])
    for ref in refs.filter(eval_mode__in=['multi', 'dual']):
        for item in ref.signal_items.exclude(consistency='single'):
            ws4.append([
                ref.title, item.signal_question,
                item.model1_id, item.model1_judgment, item.model1_reason,
                item.model2_id, item.model2_judgment, item.model2_reason,
                item.consistency, item.system_recommendation, item.human_judgment,
            ])

    # 返回文件流
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'qa_export_{project.name}_{quality_method}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    return filename, buf.getvalue()
