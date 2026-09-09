"""初筛结果的 RIS 和 Excel 导出契约。"""

import json
import tempfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import openpyxl
from django.test import TestCase

from core.artifacts.types import ArtifactType
from core.models import DataFile, Project
from core.services.project_service import initialize_project
from core.screening.executors.export_handler import ExportHandler


FIXTURES = Path(__file__).parent / 'fixtures'


class ScreeningExportGoldenTests(TestCase):
    def make_handler(self, workspace):
        executor = SimpleNamespace(
            logger=MagicMock(),
            workspace=Path(workspace),
            project_obj=MagicMock(),
            task_obj=None,
            step_obj=MagicMock(),
            stage_obj=MagicMock(),
            project_id=1,
            config={},
        )
        return ExportHandler(executor)

    def test_ris_output_matches_golden_file(self):
        result = {
            'title': 'Golden Study',
            'authors': ['Zhang, San', 'Li, Si'],
            'year': '2024',
            'journal': 'Evidence Journal',
            'volume': '12',
            'issue': '3',
            'page': '101-109',
            'doi': '10.1000/golden',
            'abstract': 'Golden abstract',
            'url': 'https://example.org/golden',
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            handler = self.make_handler(temp_dir)
            path = handler._generate_ris([result], 'golden', 'fixed')
            actual = path.read_text(encoding='utf-8')
        expected = (FIXTURES / 'golden' / 'screening_included.ris').read_text(encoding='utf-8')
        self.assertEqual(actual, expected)

    def test_excel_rows_match_semantic_golden(self):
        results = [
            {
                'source_xml': 'excluded.xml',
                'title': 'Excluded Study',
                'include_or_not': 'yes',
            },
            {
                'source_xml': 'included.xml',
                'title': 'Included Study',
                'decision': 'included',
                'exclusion_reason': 'must be cleared',
                'number_exclusion_reason': '9',
            },
        ]
        manual_reviews = {
            'excluded.xml': SimpleNamespace(
                decision='excluded',
                reason='Wrong population',
                is_override=True,
            )
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            handler = self.make_handler(temp_dir)
            with patch.object(handler, '_load_xml_fields', return_value={}), patch.object(
                handler, '_load_extraction_field_names', return_value=[]
            ):
                path = handler._generate_excel(
                    results,
                    'all',
                    'golden',
                    'fixed',
                    manual_reviews,
                    ['Adults only', 'Wrong population'],
                )
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = list(next(rows))
            selected = []
            keys = ['include_or_not', 'manual_override', 'exclusion_reason_id', 'exclusion_reason', 'Title']
            for row in rows:
                record = dict(zip(headers, row))
                selected.append({key: record[key] for key in keys})
            workbook.close()

        expected = json.loads((FIXTURES / 'golden' / 'screening_excel_rows.json').read_text(encoding='utf-8'))
        self.assertEqual(selected, expected)

    def test_excel_marks_waived_conflict_in_existing_columns(self):
        result = {
            'source_xml': 'conflict.xml',
            'title': 'Conflicting Study',
            'decision': 'excluded',
            'consensus': 'conflict',
            'multi_model_results': [
                {
                    'model_name': 'Model Include',
                    'decision': 'included',
                    'reason': '符合纳入标准',
                },
                {
                    'model_name': 'Model Exclude',
                    'decision': 'excluded',
                    'reason_id': '2',
                    'reason': '研究人群不符',
                },
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            handler = self.make_handler(temp_dir)
            with patch.object(handler, '_load_xml_fields', return_value={}), patch.object(
                handler, '_load_extraction_field_names', return_value=[]
            ):
                path = handler._generate_excel(
                    [result], 'all', 'golden', 'conflict', {}, [],
                )
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = workbook.active.iter_rows(values_only=True)
            headers = list(next(rows))
            record = dict(zip(headers, next(rows)))
            workbook.close()

        self.assertEqual(record['include_or_not'], 'conflict')
        self.assertIsNone(record['exclusion_reason_id'])
        self.assertIn('AI模型存在分歧', record['exclusion_reason'])
        self.assertIn('已豁免导出', record['exclusion_reason'])
        self.assertIn('Model Include：纳入；理由：符合纳入标准', record['exclusion_reason'])
        self.assertIn('Model Exclude：排除（排除标准 2）；理由：研究人群不符', record['exclusion_reason'])

    def test_ris_marks_waived_conflict_as_note(self):
        result = {
            'title': 'Conflicting RIS Study',
            '_export_final_decision': 'conflict',
            'multi_model_results': [
                {'model_name': 'Model A', 'decision': 'included', 'reason': '符合标准'},
                {'model_name': 'Model B', 'decision': 'excluded', 'reason': '人群不符'},
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            handler = self.make_handler(temp_dir)
            path = handler._generate_ris([result], 'golden', 'conflict')
            content = path.read_text(encoding='utf-8')

        self.assertIn('TI  - Conflicting RIS Study', content)
        self.assertIn('N1  - AI模型存在分歧', content)
        self.assertIn('Model A：纳入；理由：符合标准', content)
        self.assertIn('Model B：排除；理由：人群不符', content)

    def test_ris_only_conflict_is_not_written_to_included_excel(self):
        result = {
            'title': 'RIS Only Conflict',
            'consensus': 'conflict',
            '_export_final_decision': 'conflict',
            '_export_include_excel': False,
            '_export_xml_fields': {},
        }
        callback = MagicMock()
        with tempfile.TemporaryDirectory() as temp_dir:
            handler = self.make_handler(temp_dir)
            with patch.object(handler, '_load_extraction_field_names', return_value=[]):
                path = handler._generate_excel(
                    [result], 'included', 'golden', 'ris-only', {}, [],
                    on_record=callback,
                )
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(len(rows), 1)  # 仅表头
        callback.assert_called_once()


class ScreeningExportBatchingTests(TestCase):
    def setUp(self):
        from django.contrib.auth import get_user_model

        self.user = get_user_model().objects.create_user('export-user', password='pw')
        self.project = Project.objects.create(name='流式导出项目', owner=self.user)
        initialize_project(self.project, self.user)
        self.ai_step = self.project.stages.get(stage_key='SCREEN_1').steps.get(step_key='ai_screen')

    def test_prepared_results_are_loaded_in_bounded_batches(self):
        for index in range(5):
            DataFile.objects.create(
                project=self.project,
                stage=self.ai_step.stage,
                step=self.ai_step,
                filename=f'result-{index}.json',
                file='',
                data_category='output',
                source='tool_generated',
                metadata={
                    'artifact_type': ArtifactType.SCREENING_RESULT_JSON,
                    'source_xml': f'{index}.xml',
                    'decision': 'included' if index % 2 == 0 else 'excluded',
                    'consensus': 'included' if index % 2 == 0 else 'excluded',
                },
                created_by=self.user,
            )

        executor = SimpleNamespace(
            logger=MagicMock(), workspace=Path('.'), project_obj=self.project,
            task_obj=None, step_obj=MagicMock(), stage_obj=MagicMock(),
            project_id=self.project.id, config={}, check_stop_signal=lambda: False,
        )
        handler = ExportHandler(executor)
        result_files = DataFile.objects.filter(step=self.ai_step)
        stats = {'total': 0, 'included': 0, 'excluded': 0, 'manual_overrides': 0, 'exported': 0}

        with patch(
            'core.screening.selectors.load_ai_result_file',
            side_effect=lambda data_file: dict(data_file.metadata),
        ), patch(
            'core.screening.selectors.load_xml_fields_bulk', return_value={},
        ) as load_xml:
            results = list(handler._iter_prepared_results(
                result_files, 'all', stats, batch_size=2,
            ))

        self.assertEqual(len(results), 5)
        self.assertEqual([len(call.args[0]) for call in load_xml.call_args_list], [2, 2, 1])
        self.assertEqual(stats['total'], 5)
        self.assertEqual(stats['included'], 3)
        self.assertEqual(stats['excluded'], 2)

    def test_unresolved_conflict_requires_explicit_waiver(self):
        conflict = DataFile.objects.create(
            project=self.project,
            stage=self.ai_step.stage,
            step=self.ai_step,
            filename='conflict.json',
            file='',
            data_category='output',
            source='tool_generated',
            metadata={
                'artifact_type': ArtifactType.SCREENING_RESULT_JSON,
                'source_xml': 'conflict.xml',
                'decision': 'excluded',
                'consensus': 'conflict',
            },
            created_by=self.user,
        )
        executor = SimpleNamespace(
            logger=MagicMock(), workspace=Path('.'), project_obj=self.project,
            task_obj=None, step_obj=MagicMock(), stage_obj=MagicMock(),
            project_id=self.project.id, config={}, check_stop_signal=lambda: False,
        )
        handler = ExportHandler(executor)
        stats = {'total': 0, 'included': 0, 'excluded': 0, 'manual_overrides': 0, 'exported': 0}

        patches = (
            patch('core.screening.selectors.load_ai_result_file', return_value=dict(conflict.metadata)),
            patch('core.screening.selectors.load_xml_fields_bulk', return_value={}),
        )
        with patches[0], patches[1]:
            with self.assertRaisesRegex(RuntimeError, 'AI 分歧'):
                list(handler._iter_prepared_results(
                    DataFile.objects.filter(pk=conflict.pk), 'all', stats,
                ))

        handler.config = {'allow_unresolved_conflicts': True}
        stats = {'total': 0, 'included': 0, 'excluded': 0, 'manual_overrides': 0, 'exported': 0}
        with patch(
            'core.screening.selectors.load_ai_result_file', return_value=dict(conflict.metadata),
        ), patch('core.screening.selectors.load_xml_fields_bulk', return_value={}):
            results = list(handler._iter_prepared_results(
                DataFile.objects.filter(pk=conflict.pk), 'all', stats,
            ))

        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]['_export_final_decision'], 'conflict')
