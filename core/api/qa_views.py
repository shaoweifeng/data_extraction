"""
文献质量评价 API

接口列表（前缀 /api/qa/）：
  GET  /refs/                     → 获取项目文献列表
  POST /refs/import/              → 从初筛/复筛导入
  POST /refs/upload/              → 上传全文文件创建文献
  PATCH /refs/<id>/               → 更新单篇文献（方法选择等）
  POST /refs/batch-method/        → 批量设置方法
  GET  /methods/                  → 获取所有可用质量评价方法

  POST /eval/start/               → 启动 AI 质量评价任务
  GET  /eval/progress/            → 查询评价进度

  GET  /signal-items/             → 获取文献信号问题列表
  PATCH /signal-items/<id>/confirm/   → 确认单条信号问题
  POST /signal-items/batch-confirm/   → 一键批量确认

  GET  /domain-results/           → 获取领域汇总结果

  POST /chart/generate/           → 触发图表生成（前端渲染数据）
  GET  /chart/                    → 获取图表信息与数据
  POST /export/excel/             → 生成并下载 Excel
  GET  /export/status/            → 查询导出状态
"""

import io
import os
import csv
import base64
import json
import logging
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone as dt_tz

import matplotlib
matplotlib.use('Agg')   # 非交互后端，避免 GUI 依赖
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Circle
import numpy as np
import pandas as pd

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db import transaction, models
from django.utils import timezone

from core.models import (
    Project, DataFile, QAReference, QASignalItem, QADomainResult, QAChart
)
from core.services.quality_methods import get_method_config, get_all_methods_meta, AI_SUPPORTED_METHODS

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# 工具
# ─────────────────────────────────────────────────────────────────────────────

def _json_ok(data=None, status=200):
    return JsonResponse({'ok': True, 'data': data or {}}, status=status)


def _json_err(msg, status=400):
    return JsonResponse({'ok': False, 'error': msg}, status=status)


def _get_project(request, project_id):
    """获取项目，校验归属（权限从宽：任何登录用户可访问其参与项目）"""
    try:
        return Project.objects.get(pk=project_id)
    except Project.DoesNotExist:
        return None


def _serialize_ref(ref: QAReference) -> dict:
    return {
        'id':             ref.id,
        'title':          ref.title,
        'first_author':   ref.first_author,
        'year':           ref.year,
        'journal':        ref.journal,
        'abstract':       ref.abstract,
        'doi':            ref.doi,
        'source_type':    ref.source_type,
        'fulltext_status':ref.fulltext_status,
        'fulltext_file_id': ref.fulltext_file_id,
        'fulltext_url':   ref.fulltext_file.file.url if ref.fulltext_file else None,
        'quality_method': ref.quality_method,
        'eval_mode':      ref.eval_mode,
        'selected_models':ref.selected_models,
        'ai_eval_status': ref.ai_eval_status,
        'review_status':  ref.review_status,
        'created_at':     ref.created_at.isoformat(),
        'updated_at':     ref.updated_at.isoformat(),
    }


def _safe_int(val):
    """将字符串安全转换为 int，失败返回 None。"""
    try:
        return int(val) if val else None
    except (ValueError, TypeError):
        return None


def _safe_list(val) -> list:
    """确保 options 字段始终返回 list，兼容历史数据中意外存成字符串的情况。"""
    if isinstance(val, list):
        return val
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            return parsed if isinstance(parsed, list) else [parsed]
        except Exception:
            return [val] if val else []
    return []


_DOMAIN_NAME_MAP = {
    'patient_selection':  '患者选择',
    'index_test':         '待评价试验',
    'reference_standard': '参考标准',
    'flow_timing':        '流程与时间',
    'bias_risk':          '偏倚风险',
    'applicability':      '适用性',
}

def _serialize_signal(item: QASignalItem) -> dict:
    return {
        'id':               item.id,
        'qa_ref_id':        item.qa_ref_id,
        'quality_method':   item.quality_method,
        'domain':           item.domain,
        'domain_name':      _DOMAIN_NAME_MAP.get(item.domain, item.domain),
        'result_type':      item.result_type,
        'signal_key':       item.signal_key,
        'signal_question':  item.signal_question,
        'signal_description': item.signal_description,
        'options':          _safe_list(item.options),
        # 汇总字段
        'ai_judgment':      item.ai_judgment,
        'ai_reason':        item.ai_reason,
        'ai_evidence':      item.ai_evidence,
        'ai_evidence_page': item.ai_evidence_page,
        # N 模型原始结果列表
        'model_results':    item.model_results or [],
        # 向后兼容双模型字段
        'model1_id':        item.model1_id,
        'model1_judgment':  item.model1_judgment,
        'model1_reason':    item.model1_reason,
        'model2_id':        item.model2_id,
        'model2_judgment':  item.model2_judgment,
        'model2_reason':    item.model2_reason,
        'consistency':      item.consistency,
        'system_recommendation': item.system_recommendation,
        'pre_selected':     item.pre_selected,
        # 人工确认
        'human_judgment':   item.human_judgment,
        'is_modified':      item.is_modified,
        'original_ai_judgment': item.original_ai_judgment,
        'is_confirmed':     item.is_confirmed,
        'confirmed_by':     item.confirmed_by.username if item.confirmed_by else None,
        'confirmed_at':     item.confirmed_at.isoformat() if item.confirmed_at else None,
    }


def _serialize_domain(dr: QADomainResult) -> dict:
    return {
        'id':                          dr.id,
        'qa_ref_id':                   dr.qa_ref_id,
        'domain':                      dr.domain,
        'domain_name':                 dr.domain_name,
        'bias_risk_result':            dr.bias_risk_result,
        'applicability_result':        dr.applicability_result,
        'bias_all_confirmed':          dr.bias_all_confirmed,
        'applicability_all_confirmed': dr.applicability_all_confirmed,
    }


def _recalc_domain_results(qa_ref: QAReference):
    """
    根据已确认的信号问题，重新计算并更新 QADomainResult。
    规则（QUADAS-2 / NOS 通用）：
      - bias_risk: 任意一条 confirmed human_judgment 中有风险倾向 → high；
                   全部 low → low；否则 unclear；未全部确认 → pending
      - applicability: 同上逻辑
    """
    items = list(qa_ref.signal_items.all())
    method_cfg = None
    try:
        method_cfg = get_method_config(qa_ref.quality_method)
    except Exception:
        pass

    # 按 domain 分组
    domains = {}
    for item in items:
        domains.setdefault(item.domain, []).append(item)

    for domain_key, domain_items in domains.items():
        # 找领域名
        domain_name = domain_key
        if method_cfg:
            for d in method_cfg.get('domains', []):
                if d['key'] == domain_key:
                    domain_name = d['name']
                    break

        bias_items  = [i for i in domain_items if i.result_type == 'bias_risk']
        applic_items = [i for i in domain_items if i.result_type == 'applicability']

        def calc_result(signal_list):
            if not signal_list:
                return 'na', True
            confirmed = [i for i in signal_list if i.is_confirmed]
            if len(confirmed) < len(signal_list):
                return 'pending', False
            judgments = [i.human_judgment for i in confirmed]
            # QUADAS-2 bias: 有'否'→high；有'不清楚'→unclear；否则low
            # NOS: 有✗ → high, 有★★ or ★ → 计星，暂用简化逻辑
            if any(j in ('否', '✗') for j in judgments):
                return 'high', True
            if any(j in ('不清楚', '高') for j in judgments):
                return 'unclear', True
            return 'low', True

        bias_result, bias_confirmed = calc_result(bias_items)
        applic_result, applic_confirmed = calc_result(applic_items)

        QADomainResult.objects.update_or_create(
            qa_ref=qa_ref,
            domain=domain_key,
            defaults={
                'domain_name':                 domain_name,
                'bias_risk_result':            bias_result,
                'applicability_result':        applic_result,
                'bias_all_confirmed':          bias_confirmed,
                'applicability_all_confirmed': applic_confirmed,
            },
        )

    # 更新文献级 review_status
    all_items = list(qa_ref.signal_items.all())
    if all_items:
        confirmed_count = sum(1 for i in all_items if i.is_confirmed)
        if confirmed_count == 0:
            new_status = 'not_started'
        elif confirmed_count == len(all_items):
            new_status = 'confirmed'
        else:
            new_status = 'partial'
        QAReference.objects.filter(pk=qa_ref.pk).update(review_status=new_status)


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/qa/methods/
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def methods_list(request):
    """返回所有可用质量评价方法"""
    metas = get_all_methods_meta()
    return _json_ok(metas)


# ─────────────────────────────────────────────────────────────────────────────
# 文献管理
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def ref_list(request):
    """GET /api/qa/refs/?project_id="""
    project_id = request.GET.get('project_id')
    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)
    refs = QAReference.objects.filter(project=project).select_related('fulltext_file')
    return _json_ok([_serialize_ref(r) for r in refs])


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def ref_import(request):
    """POST /api/qa/refs/import/ — 从初筛/复筛导入已纳入文献"""
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    project_id  = body.get('project_id')
    source_stage = body.get('source_stage', 'SCREEN_1')   # SCREEN_1 or SCREEN_2
    ref_ids     = body.get('ref_ids', [])                  # 空 = 全部已纳入

    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    # ── 每次导入 = 清空重建：删除所有已有 QA 数据（含信号问题、领域结果，CASCADE 自动级联）
    # QAChart 挂在 project 上需单独清空
    from core.models import ManualReview, QAChart
    from core.api.review_views import _load_ai_results

    with transaction.atomic():
        # 删除所有文献（QASignalItem / QADomainResult 会 CASCADE 跟着删）
        QAReference.objects.filter(project=project).delete()
        # 删除图表缓存
        QAChart.objects.filter(project=project).delete()

        # ── 收集"最终纳入"文献 ─────────────────────────────────────────────
        # 逻辑与导出步骤完全一致：
        #   人工决定（included/excluded）优先；无人工记录则取 AI 决定
        # 同一 source_xml 在本次重建中只创建一条

        # 1. 所有 ManualReview 记录（该 stage）
        manual_reviews = {
            mr.source_xml: mr
            for mr in ManualReview.objects.filter(
                project=project,
                step__stage__stage_key=source_stage,
            )
        }

        # 2. AI 结果列表（每条是一个 dict，含 source_xml / title / decision 等）
        ai_results = _load_ai_results(project.id)

        # 3. 按导出模块同款逻辑判断最终决定
        def _final_included(r: dict) -> bool:
            source_xml  = r.get('source_xml', '')
            mr = manual_reviews.get(source_xml)
            if mr and mr.decision in ('included', 'excluded'):
                return mr.decision == 'included'
            # 无人工记录 → 取 AI 决定
            consensus = r.get('consensus') or r.get('decision', '')
            if consensus in ('included', 'excluded', 'conflict'):
                return consensus == 'included'
            v = r.get('include_or_not', '')
            if v:
                return v.lower() == 'yes'
            return False

        imported = []
        seen_xml = set()   # 防止 AI 结果文件中同一篇出现多次
        for r in ai_results:
            source_xml = r.get('source_xml', '')
            if not source_xml or source_xml in seen_xml:
                continue
            seen_xml.add(source_xml)

            if not _final_included(r):
                continue

            mr = manual_reviews.get(source_xml)
            ref = QAReference.objects.create(
                project=project,
                title=r.get('title', '') or source_xml,
                first_author=(r.get('authors') or '').split(';')[0].split(',')[0][:100],
                year=_safe_int(r.get('year', '')),
                journal=(r.get('journal') or '')[:300],
                doi=(r.get('doi') or '')[:200],
                source_type='screening_import',
                source_ref_id=mr.id if mr else None,
                fulltext_status='pending',
            )
            imported.append(ref.id)

    # ActivityLog
    from core.models import ActivityLog as _ActivityLog
    _ActivityLog.objects.create(
        project=project,
        operation_type='qa_import',
        operation_detail={'imported': len(imported), 'source_stage': source_stage},
        created_by=request.user,
    )
    return _json_ok({'imported': len(imported), 'skipped': 0, 'ref_ids': imported})


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def ref_upload(request):
    """POST /api/qa/refs/upload/ — 上传全文 PDF，自动识别文献信息"""
    project_id = request.POST.get('project_id')
    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    files = request.FILES.getlist('files')
    if not files:
        return _json_err('未上传文件')

    created_refs = []
    for f in files:
        if not f.name.lower().endswith('.pdf'):
            continue
        # 保存文件
        data_file = DataFile.objects.create(
            project=project,
            filename=f.name,
            file=f,
            source='upload',
            data_category='input',
            description='质量评价全文PDF',
            created_by=request.user,
        )
        # 尝试从 PDF 解析基础信息（简化：仅用文件名作为标题）
        title = f.name.replace('.pdf', '').replace('_', ' ')
        ref = QAReference.objects.create(
            project=project,
            title=title,
            source_type='fulltext_upload',
            fulltext_file=data_file,
            fulltext_status='available',
        )
        # 异步触发 PDF 解析提取元数据（非阻塞）
        try:
            from core.tasks import parse_qa_pdf_meta
            parse_qa_pdf_meta.delay(ref.id)
        except Exception:
            pass  # Celery 未就绪时跳过，不影响上传
        created_refs.append(_serialize_ref(ref))

    # ActivityLog
    from core.models import ActivityLog
    if created_refs:
        ActivityLog.objects.create(
            project=project,
            operation_type='qa_upload_pdf',
            operation_detail={'count': len(created_refs), 'filenames': [r['title'] for r in created_refs]},
            created_by=request.user,
        )
    return _json_ok({'created': len(created_refs), 'refs': created_refs}, status=201)


@csrf_exempt
@login_required
@require_http_methods(['PATCH'])
def ref_update(request, ref_id):
    """PATCH /api/qa/refs/<id>/ — 更新单篇文献（方法选择等）"""
    try:
        ref = QAReference.objects.get(pk=ref_id)
    except QAReference.DoesNotExist:
        return _json_err('文献不存在', 404)

    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    updatable = ['quality_method', 'eval_mode', 'selected_models', 'fulltext_status', 'title', 'first_author', 'year', 'journal']
    changed = False
    for field in updatable:
        if field in body:
            setattr(ref, field, body[field])
            changed = True

    # 绑定全文
    if 'fulltext_file_id' in body:
        try:
            df = DataFile.objects.get(pk=body['fulltext_file_id'], project=ref.project)
            ref.fulltext_file = df
            ref.fulltext_status = 'available'
            changed = True
        except DataFile.DoesNotExist:
            return _json_err('文件不存在或不属于该项目')

    if changed:
        ref.save()

    return _json_ok(_serialize_ref(ref))


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def ref_batch_method(request):
    """POST /api/qa/refs/batch-method/ — 批量设置质量评价方法"""
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    ref_ids       = body.get('ref_ids', [])
    quality_method = body.get('quality_method', '')
    if not ref_ids or not quality_method:
        return _json_err('缺少 ref_ids 或 quality_method')

    updated = QAReference.objects.filter(pk__in=ref_ids).update(quality_method=quality_method)

    # ActivityLog
    from core.models import ActivityLog
    if updated:
        # 取项目 id（从第一条 ref）
        first_ref = QAReference.objects.filter(pk__in=ref_ids).first()
        if first_ref:
            ActivityLog.objects.create(
                project=first_ref.project,
                operation_type='qa_set_method',
                operation_detail={'method': quality_method, 'count': updated},
                created_by=request.user,
            )
    return _json_ok({'updated': updated})


# ─────────────────────────────────────────────────────────────────────────────
# AI 评价
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@login_required
@require_http_methods(['POST'])
def eval_start(request):
    """POST /api/qa/eval/start/ — 启动 AI 质量评价任务（接入 Task/ActivityLog 体系）"""
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    project_id = body.get('project_id')
    ref_ids    = body.get('ref_ids', [])
    model_ids  = body.get('model_ids', [])
    # eval_mode 由服务端根据 model_ids 长度自动推断，前端不再需要传
    eval_mode  = body.get('eval_mode', '')  # 保留向后兼容，忽略实际值

    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    # 积分校验
    from core.services.billing_service import get_balance, estimate_credits
    balance = get_balance(request.user)
    qs = QAReference.objects.filter(project=project)
    if ref_ids:
        qs = qs.filter(pk__in=ref_ids)
    qs = qs.filter(quality_method__in=AI_SUPPORTED_METHODS).exclude(quality_method='')
    evaluable_count = qs.count()

    if evaluable_count == 0:
        return _json_err('没有可评价的文献（请先为文献选择支持 AI 评价的质量评价方法）')

    estimated = estimate_credits(evaluable_count, model_ids)
    if balance < estimated:
        return _json_err(f'积分不足（当前 {balance}，需要约 {estimated}）')

    ref_ids_to_eval = list(qs.values_list('pk', flat=True))

    # ── 走统一的 TaskScheduler 体系 ──────────────────────────────────────
    try:
        from core.scheduler import TaskScheduler
        from core.services.task_service import log_task_start
        from core.models import ActivityLog

        scheduler = TaskScheduler(project.id)
        task = scheduler.start_step(
            'qa_eval',
            request.user.id,
            ref_ids=ref_ids_to_eval,
            model_ids=model_ids,
        )
        task_id = task.id

        # ActivityLog
        log_task_start(project.id, 'qa_eval', task_id, request.user)

    except Exception as e:
        logger.error(f'[qa_eval] TaskScheduler 提交失败: {e}')
        return _json_err(f'任务提交失败: {e}')

    return _json_ok({
        'task_id':          task_id,
        'evaluable_count':  evaluable_count,
        'ref_ids':          ref_ids_to_eval,
        'estimated_credits': estimated,
    })


def _get_qa_token_stats(project, user):
    """从 TokenUsageLog 查本项目最近一次 QA 评价的实际 token/积分消耗。"""
    try:
        from core.models_billing import TokenUsageLog
        # 优先通过 project 字段直接关联（新写法），兼容通过 task 关联的旧记录
        log = (
            TokenUsageLog.objects
            .filter(user=user)
            .filter(
                models.Q(project=project) | models.Q(task__project=project)
            )
            .order_by('-created_at')
            .first()
        )
        if log:
            return {
                'total_tokens':      log.total_tokens,
                'prompt_tokens':     log.prompt_tokens,
                'completion_tokens': log.completion_tokens,
                'credits_consumed':  log.credits_consumed,
                'ref_count':         log.ref_count,
                'model':             log.model,
                'recorded_at':       log.created_at.isoformat(),
            }
    except Exception:
        pass
    return None


@login_required
@require_http_methods(['GET'])
def eval_progress(request):
    """GET /api/qa/eval/progress/?project_id="""
    project_id = request.GET.get('project_id')
    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    refs = QAReference.objects.filter(project=project)
    total       = refs.count()
    pending     = refs.filter(ai_eval_status='pending').count()
    running     = refs.filter(ai_eval_status='running').count()
    completed   = refs.filter(ai_eval_status='completed').count()
    failed      = refs.filter(ai_eval_status__in=['failed', 'skipped_no_fulltext', 'skipped_no_method']).count()
    abstract_only = refs.filter(ai_eval_status='abstract_only').count()

    # 双模型信号问题一致性统计
    divergent_count = QASignalItem.objects.filter(
        qa_ref__project=project,
        consistency='divergent',
    ).count()

    unconfirmed_count = QASignalItem.objects.filter(
        qa_ref__project=project,
        qa_ref__ai_eval_status__in=['completed', 'abstract_only'],
        is_confirmed=False,
    ).count()

    # 文献级进度列表
    ref_list_data = []
    for ref in refs.select_related('fulltext_file'):
        # 双模型分歧数
        div = QASignalItem.objects.filter(qa_ref=ref, consistency='divergent').count() if ref.eval_mode == 'dual' else 0
        ref_list_data.append({
            'id':             ref.id,
            'title':          ref.title,
            'quality_method': ref.quality_method,
            'eval_mode':      ref.eval_mode,
            'ai_eval_status': ref.ai_eval_status,
            'review_status':  ref.review_status,
            'divergent_count': div,
        })

    return _json_ok({
        'summary': {
            'total':        total,
            'pending':      pending,
            'running':      running,
            'completed':    completed + abstract_only,
            'failed':       failed,
            'abstract_only': abstract_only,
            'divergent_signal_count': divergent_count,
        },
        'refs': ref_list_data,
        'token_stats': _get_qa_token_stats(project, request.user),
    })


# ─────────────────────────────────────────────────────────────────────────────
# 信号问题
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def signal_items_list(request):
    """GET /api/qa/signal-items/?qa_ref_id=&domain=&result_type=&is_confirmed="""
    qa_ref_id = request.GET.get('qa_ref_id')
    if not qa_ref_id:
        return _json_err('缺少 qa_ref_id')
    try:
        ref = QAReference.objects.get(pk=qa_ref_id)
    except QAReference.DoesNotExist:
        return _json_err('文献不存在', 404)

    qs = ref.signal_items.select_related('confirmed_by').all()
    if request.GET.get('domain'):
        qs = qs.filter(domain=request.GET['domain'])
    if request.GET.get('result_type'):
        qs = qs.filter(result_type=request.GET['result_type'])
    if request.GET.get('is_confirmed') == 'false':
        qs = qs.filter(is_confirmed=False)
    elif request.GET.get('is_confirmed') == 'true':
        qs = qs.filter(is_confirmed=True)

    return _json_ok([_serialize_signal(i) for i in qs])


@csrf_exempt
@login_required
@require_http_methods(['PATCH'])
def signal_item_confirm(request, item_id):
    """PATCH /api/qa/signal-items/<id>/confirm/ — 确认单条信号问题"""
    try:
        item = QASignalItem.objects.select_related('qa_ref').get(pk=item_id)
    except QASignalItem.DoesNotExist:
        return _json_err('信号问题不存在', 404)
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    human_judgment = body.get('human_judgment', '').strip()
    if not human_judgment:
        return _json_err('缺少 human_judgment')

    with transaction.atomic():
        # 记录修改
        if item.ai_judgment and human_judgment != item.ai_judgment and not item.is_modified:
            item.is_modified = True
            item.original_ai_judgment = item.ai_judgment
        # 也兼容双模型推荐
        if item.system_recommendation and human_judgment != item.system_recommendation and not item.is_modified:
            item.is_modified = True
            item.original_ai_judgment = item.system_recommendation

        item.human_judgment = human_judgment
        item.is_confirmed  = True
        item.confirmed_by  = request.user
        item.confirmed_at  = timezone.now()
        item.save()

        # 重新聚合领域结果
        _recalc_domain_results(item.qa_ref)

    # ActivityLog
    from core.models import ActivityLog
    ActivityLog.objects.create(
        project=item.qa_ref.project,
        operation_type='qa_confirm_signal',
        operation_detail={
            'qa_ref_id': item.qa_ref_id,
            'signal_key': item.signal_key,
            'judgment': human_judgment,
            'modified': item.is_modified,
        },
        created_by=request.user,
    )
    return _json_ok(_serialize_signal(item))


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def signal_batch_confirm(request):
    """
    POST /api/qa/signal-items/batch-confirm/
    Body: {
      "qa_ref_id": 1,
      "confirm_mode": "adopt_preselected" | "adopt_ai" | "specific_keys",
      "signal_keys": [...],   # confirm_mode=specific_keys 时使用
    }
    """
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    qa_ref_id    = body.get('qa_ref_id')
    confirm_mode = body.get('confirm_mode', 'adopt_preselected')
    signal_keys  = body.get('signal_keys', [])

    if not qa_ref_id:
        return _json_err('缺少 qa_ref_id')
    try:
        ref = QAReference.objects.get(pk=qa_ref_id)
    except QAReference.DoesNotExist:
        return _json_err('文献不存在', 404)

    items = ref.signal_items.all()
    if confirm_mode == 'specific_keys' and signal_keys:
        items = items.filter(signal_key__in=signal_keys)

    confirmed_count = 0
    now = timezone.now()
    with transaction.atomic():
        for item in items:
            if confirm_mode == 'adopt_preselected':
                if not item.pre_selected:
                    continue
                judgment = item.pre_selected
            elif confirm_mode == 'adopt_ai':
                if not item.ai_judgment:
                    continue
                judgment = item.ai_judgment
            else:
                # specific_keys: 使用 pre_selected 或 ai_judgment
                judgment = item.pre_selected or item.ai_judgment
                if not judgment:
                    continue

            if item.ai_judgment and judgment != item.ai_judgment and not item.is_modified:
                item.is_modified = True
                item.original_ai_judgment = item.ai_judgment

            item.human_judgment = judgment
            item.is_confirmed  = True
            item.confirmed_by  = request.user
            item.confirmed_at  = now
            item.save(update_fields=[
                'human_judgment', 'is_confirmed', 'confirmed_by', 'confirmed_at',
                'is_modified', 'original_ai_judgment',
            ])
            confirmed_count += 1

        # 重新聚合
        _recalc_domain_results(ref)

    # ActivityLog
    from core.models import ActivityLog
    ActivityLog.objects.create(
        project=ref.project,
        operation_type='qa_batch_confirm',
        operation_detail={
            'qa_ref_id': qa_ref_id,
            'confirm_mode': confirm_mode,
            'confirmed_count': confirmed_count,
        },
        created_by=request.user,
    )
    return _json_ok({'confirmed': confirmed_count})


# ─────────────────────────────────────────────────────────────────────────────
# 领域汇总结果
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def domain_results(request):
    """GET /api/qa/domain-results/?qa_ref_id="""
    qa_ref_id = request.GET.get('qa_ref_id')
    if not qa_ref_id:
        return _json_err('缺少 qa_ref_id')
    qs = QADomainResult.objects.filter(qa_ref_id=qa_ref_id)
    return _json_ok([_serialize_domain(dr) for dr in qs])


# ─────────────────────────────────────────────────────────────────────────────
# 图表
# ─────────────────────────────────────────────────────────────────────────────

def _build_chart_data(project, quality_method, ref_ids=None):
    """
    计算图表所需的数据结构（不渲染图片）。
    返回 (traffic_light_data, proportion_data, bias_domains, applic_domains, method_cfg)
    """
    from core.services.quality_methods import get_method_config
    method_cfg = get_method_config(quality_method)
    domains        = method_cfg['domains']
    bias_domains   = [d for d in domains if d['has_bias_risk']]
    applic_domains = [d for d in domains if d['has_applicability']]

    qs = QAReference.objects.filter(project=project, quality_method=quality_method)
    if ref_ids:
        qs = qs.filter(pk__in=ref_ids)

    traffic_light_data = []
    for ref in qs:
        domain_map = {dr.domain: dr for dr in ref.domain_results.all()}
        row = {
            'ref_id': ref.id, 'title': ref.title,
            'first_author': ref.first_author, 'year': ref.year,
            'review_status': ref.review_status,
            'bias_risk': {}, 'applicability': {},
        }
        for d in bias_domains:
            dr = domain_map.get(d['key'])
            row['bias_risk'][d['key']] = dr.bias_risk_result if dr else 'pending'
        for d in applic_domains:
            dr = domain_map.get(d['key'])
            row['applicability'][d['key']] = dr.applicability_result if dr else 'pending'
        traffic_light_data.append(row)

    proportion_data = {}
    for d in bias_domains + applic_domains:
        k = d['key']
        confirmed_refs = [r for r in traffic_light_data if r['review_status'] == 'confirmed']
        counts = {'low': 0, 'high': 0, 'unclear': 0, 'pending': 0}
        for r in confirmed_refs:
            val = r['bias_risk'].get(k) or r['applicability'].get(k) or 'pending'
            if val in counts:
                counts[val] += 1
        total = max(1, len(confirmed_refs))
        proportion_data[k] = {
            'domain_name': d['name'],
            'result_type': 'bias_risk' if d['has_bias_risk'] else 'applicability',
            'counts': counts,
            'percentages': {k2: round(v / total * 100, 1) for k2, v in counts.items()},
        }

    return traffic_light_data, proportion_data, bias_domains, applic_domains, method_cfg


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def chart_preview(request):
    """
    POST /api/qa/chart/preview/
    快速返回前端渲染所需的数据结构，不生成 PNG 图片。
    用于页面加载时的实时预览，用户编辑文献名后点「生成图片」才真正渲染 PNG。
    """
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    project_id     = body.get('project_id')
    quality_method = body.get('quality_method', 'QUADAS2')
    ref_ids        = body.get('ref_ids', [])

    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    try:
        tl, prop, bias_d, applic_d, method_cfg = _build_chart_data(project, quality_method, ref_ids or None)
    except Exception as e:
        return _json_err(f'数据构建失败: {e}')

    return _json_ok({
        'quality_method':    quality_method,
        'method_name':       method_cfg['name'],
        'bias_domains':      bias_d,
        'applic_domains':    applic_d,
        'traffic_light':     tl,
        'proportion':        prop,
        'generated_at':      None,
        'unconfirmed_count': sum(1 for r in tl if r['review_status'] != 'confirmed'),
    })


@csrf_exempt
@login_required
@require_http_methods(['POST'])
def chart_generate(request):
    """
    POST /api/qa/chart/generate/
    使用用户最终确认的文献名生成 PNG 图片（base64）并返回。
    需先调用 /preview/ 获取数据、编辑文献名后再调用本接口生成图片文件。
    """
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    project_id     = body.get('project_id')
    quality_method = body.get('quality_method', 'QUADAS2')
    ref_ids        = body.get('ref_ids', [])
    study_labels   = body.get('study_labels') or {}
    orientation    = body.get('orientation', 'horizontal')

    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    try:
        traffic_light_data, proportion_data, bias_domains, applic_domains, method_cfg = \
            _build_chart_data(project, quality_method, ref_ids or None)
    except Exception as e:
        return _json_err(f'数据构建失败: {e}')

    # 保存/更新图表记录
    chart, _ = QAChart.objects.update_or_create(
        project=project,
        quality_method=quality_method,
        defaults={
            'chart_types': ['traffic_light', 'proportion', 'detail'],
            'ref_ids':     [r['ref_id'] for r in traffic_light_data],
            'generated_at': timezone.now(),
        },
    )

    # ── 生成图表图片（base64 PNG）────────────────────────────────────────────
    traffic_b64  = _render_traffic_light(
        traffic_light_data, bias_domains, applic_domains,
        method_cfg['name'], quality_method=quality_method,
        study_labels=study_labels, orientation=orientation,
    )
    proportion_b64 = _render_proportion(
        proportion_data, method_cfg['name'],
        quality_method=quality_method,
        traffic_light_data=traffic_light_data,
        bias_domains=bias_domains,
        applic_domains=applic_domains,
        study_labels=study_labels,
    )

    from core.models import ActivityLog
    ActivityLog.objects.create(
        project=project,
        operation_type='qa_generate_chart',
        operation_detail={
            'quality_method': quality_method,
            'ref_count': len(traffic_light_data),
        },
        created_by=request.user,
    )
    return _json_ok({
        'chart_id':             chart.id,
        'quality_method':       quality_method,
        'method_name':          method_cfg['name'],
        'bias_domains':         bias_domains,
        'applic_domains':       applic_domains,
        'traffic_light':        traffic_light_data,
        'proportion':           proportion_data,
        'generated_at':         chart.generated_at.isoformat(),
        'unconfirmed_count':    sum(1 for r in traffic_light_data if r['review_status'] != 'confirmed'),
        'traffic_light_image':  traffic_b64,
        'proportion_image':     proportion_b64,
    })


# ── 图表绘制工具函数 ──────────────────────────────────────────────────────────

_RISK_COLORS = {
    'low':     '#059669',
    'high':    '#dc2626',
    'unclear': '#d97706',
    'pending': '#e2e8f0',
    'na':      '#94a3b8',
}
_RISK_MARKERS = {
    'low':     '✓',
    'high':    '✗',
    'unclear': '?',
    'pending': '○',
    'na':      '−',
}


def _fig_to_b64(fig) -> str:
    """将 matplotlib Figure 转为 data URL 格式的 base64 PNG。"""
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return 'data:image/png;base64,' + base64.b64encode(buf.read()).decode()


# macOS/Linux 中文字体优先顺序（matplotlib 降级时用）
_CJK_FONT_CANDIDATES = [
    'PingFang SC', 'PingFang HK', 'STHeiti', 'Heiti TC',
    'Hiragino Sans GB', 'WenQuanYi Micro Hei', 'SimHei', 'Microsoft YaHei',
]

def _setup_cjk_font():
    """检测并设置第一个可用的 CJK 字体，解决中文乱码。"""
    import matplotlib.font_manager as fm
    available = {f.name for f in fm.fontManager.ttflist}
    for name in _CJK_FONT_CANDIDATES:
        if name in available:
            plt.rcParams['font.family'] = name
            plt.rcParams['axes.unicode_minus'] = False
            return
    plt.rcParams['axes.unicode_minus'] = False


# ── robvis 相关 ────────────────────────────────────────────────────────────────

# 我们的 method key → robvis tool 参数的映射（不在此表里的方法走 matplotlib 降级）
_ROBVIS_TOOL_MAP = {
    'QUADAS2':  'QUADAS-2',
    'ROB2':     'ROB2',
    'ROBINS_I': 'ROBINS-I',
}

# 各 robvis 工具的判断值映射：我们内部 key → robvis 期望的英文字符串
_ROBVIS_JUDGMENT_MAP = {
    # QUADAS-2 / ROB2 / ROBINS-I 共用低/高/不清楚
    'low':     'Low',
    'high':    'High',
    'unclear': 'Some concerns',
    'pending': 'Some concerns',
    'na':      'Low',
    # ROBINS-I 额外值
    'critical':  'Critical',
    'moderate':  'Moderate',
    'serious':   'Serious',
}

_RSCRIPT = shutil.which('Rscript') or '/opt/homebrew/bin/Rscript'


def _robvis_render(csv_path: str, tool: str, out_png: str, chart_type: str = 'traffic_light') -> bool:
    """
    调用 Rscript 生成 robvis 图，写到 out_png。
    chart_type: 'traffic_light' | 'summary'
    返回是否成功。
    """
    func = 'rob_traffic_light' if chart_type == 'traffic_light' else 'rob_summary'
    r_code = f"""
suppressMessages({{
  library(robvis)
  library(ggplot2)
  dat <- read.csv('{csv_path}', stringsAsFactors=FALSE)
  p <- {func}(dat, tool='{tool}')
  ggsave('{out_png}', plot=p, width=10, height=max(4, nrow(dat)*0.55 + 2), dpi=150)
}})
"""
    try:
        result = subprocess.run(
            [_RSCRIPT, '--vanilla', '-e', r_code],
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode != 0:
            logger.warning('robvis R error: %s', result.stderr[-500:])
            return False
        return os.path.exists(out_png) and os.path.getsize(out_png) > 0
    except Exception as e:
        logger.warning('robvis subprocess error: %s', e)
        return False


def _robvis_render_quadas_combined(
        bias_csv: str, applic_csv: str, out_png: str,
        n_bias: int, n_applic: int, n_refs: int) -> bool:
    """
    QUADAS-2 专用：用 patchwork 把 bias 图和 applicability 图横向拼合。
    bias_csv   — 4列 bias 域的 CSV（标准 QUADAS-2 格式）
    applic_csv — applic 域 CSV（域数 ≤4，不足4列用 Low 占位）
    """
    w_bias   = max(5, n_bias   * 1.2 + 3)
    w_applic = max(3, n_applic * 1.2 + 1)
    fig_h    = max(4, n_refs * 0.55 + 2)
    r_code = f"""
suppressMessages({{
  library(robvis)
  library(ggplot2)
  library(patchwork)

  dat_bias   <- read.csv('{bias_csv}',   stringsAsFactors=FALSE)
  dat_applic <- read.csv('{applic_csv}', stringsAsFactors=FALSE)

  p_bias   <- rob_traffic_light(dat_bias,   tool='QUADAS-2')
  p_applic <- rob_traffic_light(dat_applic, tool='QUADAS-2')

  # 去掉右侧图的 Study 轴标签（文献名已在左侧图显示）
  p_applic2 <- p_applic +
    ggplot2::labs(title='Applicability concerns') +
    ggplot2::theme(
      axis.text.y  = ggplot2::element_blank(),
      axis.title.y = ggplot2::element_blank(),
      plot.title   = ggplot2::element_text(size=9, hjust=0.5)
    )

  p_bias2 <- p_bias +
    ggplot2::labs(title='Risk of Bias') +
    ggplot2::theme(plot.title = ggplot2::element_text(size=9, hjust=0.5))

  combined <- p_bias2 + p_applic2 +
    plot_layout(widths=c({w_bias}, {w_applic}), guides='collect') &
    ggplot2::theme(legend.position='bottom')

  ggsave('{out_png}', plot=combined,
         width={w_bias + w_applic}, height={fig_h}, dpi=150)
}})
"""
    try:
        result = subprocess.run(
            [_RSCRIPT, '--vanilla', '-e', r_code],
            capture_output=True, text=True, timeout=90,
        )
        if result.returncode != 0:
            logger.warning('robvis combined R error: %s', result.stderr[-800:])
            return False
        return os.path.exists(out_png) and os.path.getsize(out_png) > 0
    except Exception as e:
        logger.warning('robvis combined subprocess error: %s', e)
        return False


def _build_robvis_csv(traffic_light_data: list, bias_domains: list, applic_domains: list,
                      tool: str, csv_path: str, study_labels: dict = None):
    """
    把内部数据结构转为 robvis 期望的 CSV 格式。
    robvis 格式：Study, D1, D2, ..., Overall, Weight

    robvis 靠列数识别工具类型，列数必须严格匹配：
      QUADAS-2 : 4 bias 域（无 applic 列）→ 共 7 列（Study + D1-D4 + Overall + Weight）
      ROB2     : 5 bias 域                → 共 8 列
      ROBINS-I : 7 bias 域                → 共 10 列

    注意：QUADAS-2 的 applicability 域不写入 CSV，robvis 内部自己知道哪些域有适用性判断。

    study_labels: {ref_id: label_str}，前端用户自定义文献名，优先级最高
    """
    # QUADAS-2 只传 bias 域；其他方法全传（ROB2/ROBINS-I 本来就没有 applic 域）
    if tool == 'QUADAS-2':
        csv_domains = [(d['key'], 'bias') for d in bias_domains]
    else:
        csv_domains = (
            [(d['key'], 'bias') for d in bias_domains]
          + [(d['key'], 'applic') for d in applic_domains]
        )

    headers = ['Study'] + [f'D{i+1}' for i in range(len(csv_domains))] + ['Overall', 'Weight']

    def worst_judgment(row):
        vals = list(row['bias_risk'].values()) + list(row['applicability'].values())
        if 'high' in vals or 'critical' in vals:
            return 'high'
        if 'unclear' in vals or 'serious' in vals or 'moderate' in vals:
            return 'unclear'
        return 'low'

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in traffic_light_data:
            ref_id = row['ref_id']
            # 优先用前端自定义标签
            if study_labels and str(ref_id) in study_labels:
                study = study_labels[str(ref_id)][:60]
            else:
                study = (row.get('first_author') or row.get('title') or f"Ref {ref_id}")[:40]
                year = row.get('year')
                if year:
                    study = f"{study} ({year})"
            cells = [study]
            for dkey, dtype in csv_domains:
                raw = (row['bias_risk'].get(dkey) if dtype == 'bias'
                       else row['applicability'].get(dkey)) or 'unclear'
                cells.append(_ROBVIS_JUDGMENT_MAP.get(raw, 'Some concerns'))
            overall_raw = worst_judgment(row)
            cells.append(_ROBVIS_JUDGMENT_MAP.get(overall_raw, 'Some concerns'))
            cells.append(round(100 / max(1, len(traffic_light_data)), 4))
            writer.writerow(cells)


def _build_robvis_applic_csv(traffic_light_data: list, applic_domains: list,
                              csv_path: str, study_labels: dict = None):
    """
    为 QUADAS-2 applicability 专门构建 CSV。
    robvis QUADAS-2 固定需要 4 列 D，不足的用 'Low' 填充。
    """
    n_applic = len(applic_domains)
    # 最多 4 列，不足 4 列补 Low
    n_cols = 4
    headers = ['Study'] + [f'D{i+1}' for i in range(n_cols)] + ['Overall', 'Weight']

    def worst_applic(row):
        vals = list(row['applicability'].values())
        if 'high' in vals:
            return 'high'
        if 'unclear' in vals:
            return 'unclear'
        return 'low'

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in traffic_light_data:
            ref_id = row['ref_id']
            if study_labels and str(ref_id) in study_labels:
                study = study_labels[str(ref_id)][:60]
            else:
                study = (row.get('first_author') or row.get('title') or f"Ref {ref_id}")[:40]
                year = row.get('year')
                if year:
                    study = f"{study} ({year})"
            cells = [study]
            for i in range(n_cols):
                if i < n_applic:
                    dkey = applic_domains[i]['key']
                    raw  = row['applicability'].get(dkey) or 'unclear'
                    cells.append(_ROBVIS_JUDGMENT_MAP.get(raw, 'Some concerns'))
                else:
                    cells.append('Low')   # 占位
            overall_raw = worst_applic(row)
            cells.append(_ROBVIS_JUDGMENT_MAP.get(overall_raw, 'Some concerns'))
            cells.append(round(100 / max(1, len(traffic_light_data)), 4))
            writer.writerow(cells)


def _png_to_b64(png_path: str) -> str:
    """把 PNG 文件读为 base64 data URL。"""
    with open(png_path, 'rb') as f:
        return 'data:image/png;base64,' + base64.b64encode(f.read()).decode()



import matplotlib.font_manager as _fm

def _init_cjk_font():
    """用字体文件路径直接设置中文字体，绕过名称查找。
    兼容 macOS（STHeiti/Hiragino）与 Linux（fonts-noto-cjk / wqy）。
    服务器需提前安装：
        apt-get install -y fonts-noto-cjk        # Debian/Ubuntu（推荐）
      或
        yum install -y google-noto-sans-cjk-ttc  # CentOS/RHEL
    """
    candidates = [
        # macOS
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/System/Library/Fonts/STHeiti Medium.ttc',
        '/System/Library/Fonts/Hiragino Sans GB.ttc',
        '/Library/Fonts/Arial Unicode.ttf',
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
        # Linux — Noto CJK（fonts-noto-cjk）
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc',
        # Linux — WQY（wqy-microhei / wqy-zenhei）
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
        '/usr/share/fonts/wqy-microhei/wqy-microhei.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
        # Linux — 通用备用（文泉驿 CJK）
        '/usr/share/fonts/wenquanyi/wqy-microhei/wqy-microhei.ttc',
    ]
    for path in candidates:
        import os as _os
        if _os.path.exists(path):
            try:
                prop = _fm.FontProperties(fname=path)
                name = prop.get_name()
                if name not in plt.rcParams['font.sans-serif']:
                    plt.rcParams['font.sans-serif'].insert(0, name)
                    _fm.fontManager.addfont(path)
                plt.rcParams['axes.unicode_minus'] = False
                return name
            except Exception:
                continue
    plt.rcParams['axes.unicode_minus'] = False
    return None


# ── 颜色/符号（直接沿用原脚本）────────────────────────────────────────────────
_COLORS  = {"High": "#d7191c", "Unclear": "#f1e51d", "Low": "#00b83f"}
_SYMBOLS = {"High": "×", "Unclear": "?", "Low": "+"}
# 内部 key → 原脚本 key 的映射
_JUDGMENT_MAP = {
    "low":     "Low",
    "high":    "High",
    "unclear": "Unclear",
    "pending": "Unclear",
    "na":      "Low",
}


def _get_study_label(row: dict, study_labels: dict) -> str:
    ref_id = row['ref_id']
    if study_labels and str(ref_id) in study_labels:
        return study_labels[str(ref_id)][:60]
    author = row.get('first_author') or ''
    year   = row.get('year') or ''
    if author:
        return f"{author} {year}".strip()
    title = row.get('title') or f"Ref {ref_id}"
    return title[:40]


# ── 以下三个函数原封不动来自 quadas2_matplotlib_tryrun_20260830_025320.py ─────

def _draw_summary_bar(ax, summary_df, title):
    y_positions = np.arange(len(summary_df))
    left = np.zeros(len(summary_df))
    for status in ["High", "Unclear", "Low"]:
        values = summary_df[status].values
        ax.barh(y_positions, values, left=left,
                color=_COLORS[status], edgecolor="black", height=0.65, label=status)
        left += values
    ax.set_yticks(y_positions)
    ax.set_yticklabels(summary_df["domain"])
    ax.invert_yaxis()
    ax.set_xlim(0, 1)
    ax.set_xticks([0, 0.25, 0.5, 0.75, 1])
    ax.set_xticklabels(["0%", "25%", "50%", "75%", "100%"])
    ax.set_title(title, fontsize=10, fontweight="bold")
    ax.tick_params(axis="both", labelsize=8)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)


def _draw_traffic_light_matrix(ax, studies, rows, n_bias, orientation='horizontal'):
    """
    studies:     list of str
    rows:        list of {"label": str, "values": list of "High"/"Unclear"/"Low"}
    n_bias:      int — 前几行属于 Risk of Bias（其余为 Applicability Concerns）
    orientation: 'horizontal'（研究=列，默认） | 'vertical'（研究=行）
    """
    n_studies = len(studies)
    n_domains = len(rows)

    if orientation == 'vertical':
        # ── 纵向：研究=行，领域=列 ─────────────────────────────────────────
        # xlim: (-1.0 ~ n_domains+0.5)  留左侧给研究名标签
        # ylim: (-0.5 ~ n_studies+0.5)  顶部留给列头
        ax.set_xlim(-1.0, n_domains + 0.5)
        ax.set_ylim(-1.0, n_studies + 0.5)
        ax.invert_yaxis()
        ax.axis("off")

        # 领域名（列头，旋转 45°）
        for col_idx, row in enumerate(rows):
            ax.text(col_idx, -0.6, row["label"],
                    ha="center", va="bottom", rotation=45, fontsize=7)

        # bias / applic 纵向分隔线
        ax.plot([n_bias - 0.5, n_bias - 0.5],
                [-0.5, n_studies - 0.5],
                color="black", linewidth=1)

        # 圆 + 符号
        for row_idx, study in enumerate(studies):
            ax.text(-0.6, row_idx, study,
                    ha="right", va="center", fontsize=7)
            for col_idx, domain_row in enumerate(rows):
                value = domain_row["values"][row_idx]
                circle = Circle(
                    (col_idx, row_idx), radius=0.32,
                    facecolor=_COLORS[value], edgecolor="black", linewidth=0.8,
                )
                ax.add_patch(circle)
                sym_color = "black" if value == "Unclear" else "white"
                ax.text(col_idx, row_idx, _SYMBOLS[value],
                        ha="center", va="center",
                        fontsize=9, fontweight="bold", color=sym_color)

        # 底部横排组名
        ax.text((n_bias - 1) / 2, n_studies + 0.2,
                "Risk of Bias",
                ha="center", va="top", fontsize=8)
        if n_domains > n_bias:
            ax.text(n_bias + (n_domains - n_bias - 1) / 2, n_studies + 0.2,
                    "Applicability Concerns",
                    ha="center", va="top", fontsize=8)

    else:
        # ── 横向（默认）：研究=列，领域=行 ───────────────────────────────────
        n_cols = n_studies
        n_rows = n_domains
        ax.set_xlim(-0.5, n_cols + 3.5)
        ax.set_ylim(-1.5, n_rows + 0.6)
        ax.invert_yaxis()
        ax.axis("off")

        # 研究名（列头，旋转 90°）
        for i, study in enumerate(studies):
            ax.text(i, -0.8, study, ha="center", va="bottom",
                    rotation=90, fontsize=7)

        # bias / applic 水平分隔线
        ax.plot([-0.5, n_cols - 0.5],
                [n_bias - 0.5, n_bias - 0.5],
                color="black", linewidth=1)

        # 圆 + 符号 + 行标签
        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row["values"]):
                circle = Circle(
                    (col_idx, row_idx), radius=0.32,
                    facecolor=_COLORS[value], edgecolor="black", linewidth=0.8,
                )
                ax.add_patch(circle)
                sym_color = "black" if value == "Unclear" else "white"
                ax.text(col_idx, row_idx, _SYMBOLS[value],
                        ha="center", va="center",
                        fontsize=9, fontweight="bold", color=sym_color)
            ax.text(n_cols + 0.3, row_idx, row["label"],
                    ha="left", va="center", fontsize=8)

        # 右侧竖排组名
        ax.text(n_cols + 2.2, (n_bias - 1) / 2,
                "Risk of Bias",
                ha="center", va="center", rotation=270, fontsize=8)
        ax.text(n_cols + 2.2, n_bias + (n_rows - n_bias - 1) / 2,
                "Applicability Concerns",
                ha="center", va="center", rotation=270, fontsize=8)


def _draw_legend(ax):
    """用矩形色块画图例（与原脚本对齐，避免 aspect 影响）。"""
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    for x, key in zip([0.1, 0.42, 0.72], ["High", "Unclear", "Low"]):
        rect = plt.Rectangle((x, 0.25), 0.08, 0.5,
                              facecolor=_COLORS[key], edgecolor="black", linewidth=0.8)
        ax.add_patch(rect)
        ax.text(x + 0.10, 0.5, key, ha="left", va="center", fontsize=9)
    # 外框
    border = plt.Rectangle((0.05, 0.1), 0.9, 0.8,
                            facecolor="none", edgecolor="black", linewidth=1.0)
    ax.add_patch(border)


# ── 两个对外渲染接口 ───────────────────────────────────────────────────────────

def _render_traffic_light(traffic_light_data, bias_domains, applic_domains,
                          method_name, quality_method='', study_labels=None,
                          orientation='horizontal') -> str:
    """生成交通灯图（Panel B），返回 base64 data URL。"""
    _init_cjk_font()
    if not traffic_light_data:
        return None

    # 非 QUADAS2 走 robvis
    robvis_tool = _ROBVIS_TOOL_MAP.get(quality_method)
    if robvis_tool and quality_method != 'QUADAS2':
        tmpdir = tempfile.mkdtemp()
        try:
            png_path = os.path.join(tmpdir, 'tl.png')
            csv_path = os.path.join(tmpdir, 'data.csv')
            _build_robvis_csv(traffic_light_data, bias_domains, applic_domains,
                              robvis_tool, csv_path, study_labels=study_labels)
            if _robvis_render(csv_path, robvis_tool, png_path, 'traffic_light'):
                return _png_to_b64(png_path)
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    # ── 组装数据（内部 key → "High"/"Unclear"/"Low"）─────────────────────────
    studies = [_get_study_label(r, study_labels) for r in traffic_light_data]

    rows = []
    for d in bias_domains:
        rows.append({
            "label": d['name'],
            "values": [
                _JUDGMENT_MAP.get(r['bias_risk'].get(d['key'], 'pending'), 'Unclear')
                for r in traffic_light_data
            ],
        })
    for d in applic_domains:
        rows.append({
            "label": d['name'],
            "values": [
                _JUDGMENT_MAP.get(r['applicability'].get(d['key'], 'pending'), 'Unclear')
                for r in traffic_light_data
            ],
        })

    n_bias_rows = len(bias_domains)
    n_rows      = len(rows)
    n_studies   = len(studies)

    # 根据方向计算图幅
    if orientation == 'vertical':
        # 纵向：研究=行，领域=列；宽度由领域数决定，高度由研究数决定
        data_w = n_rows + 1.5      # xlim 范围
        data_h = n_studies + 1.5   # ylim 范围
        fig_w  = max(7, n_rows * 1.2 + 2)
        fig_h  = max(5, n_studies * 0.85 + 2)
    else:
        # 横向：研究=列，领域=行
        data_w = n_studies + 4.0
        data_h = n_rows + 2.1
        fig_w  = max(8, n_studies * 0.85 + 5)
        fig_h  = fig_w * data_h / data_w

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    _draw_traffic_light_matrix(ax, studies, rows, n_bias_rows, orientation=orientation)

    plt.tight_layout(pad=0.5)
    return _fig_to_b64(fig)


def _render_proportion(proportion_data, method_name, quality_method='',
                       traffic_light_data=None, bias_domains=None, applic_domains=None,
                       study_labels=None) -> str:
    """生成比例图（Panel A），返回 base64 data URL。"""
    _init_cjk_font()
    if not proportion_data:
        return None

    # 非 QUADAS2 走 robvis
    robvis_tool = _ROBVIS_TOOL_MAP.get(quality_method)
    if robvis_tool and quality_method != 'QUADAS2' and traffic_light_data and bias_domains is not None:
        tmpdir = tempfile.mkdtemp()
        try:
            csv_path = os.path.join(tmpdir, 'data.csv')
            png_path = os.path.join(tmpdir, 'summary.png')
            _build_robvis_csv(traffic_light_data, bias_domains, applic_domains or [],
                              robvis_tool, csv_path, study_labels=study_labels)
            if _robvis_render(csv_path, robvis_tool, png_path, 'summary'):
                return _png_to_b64(png_path)
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    # ── 用原脚本的 _draw_summary_bar + _draw_legend ───────────────────────────
    bias_keys   = {d['key'] for d in (bias_domains   or [])}
    applic_keys = {d['key'] for d in (applic_domains or [])}

    def _make_summary_df(domain_list, key_set):
        records = []
        for d in (domain_list or []):
            k = d['key']
            item = proportion_data.get(k) or proportion_data.get('app_' + k)
            if item is None:
                continue
            total = max(1, sum(item['counts'].values()))
            records.append({
                "domain":   d['name'],
                "High":     item['counts'].get('high', 0)   / total,
                "Unclear":  item['counts'].get('unclear', 0) / total,
                "Low":      item['counts'].get('low', 0)    / total,
            })
        return pd.DataFrame(records) if records else pd.DataFrame(
            columns=["domain", "High", "Unclear", "Low"])

    rob_summary = _make_summary_df(bias_domains,   bias_keys)
    app_summary = _make_summary_df(applic_domains, applic_keys)

    has_applic = len(app_summary) > 0

    fig = plt.figure(figsize=(11, 5.5))
    grid = fig.add_gridspec(
        nrows=2, ncols=2,
        height_ratios=[1.0, 0.22],
        width_ratios=[1, 1],
        hspace=0.5, wspace=0.35,
    )
    ax_left   = fig.add_subplot(grid[0, 0])
    ax_right  = fig.add_subplot(grid[0, 1])
    ax_legend = fig.add_subplot(grid[1, :])

    _draw_summary_bar(ax_left,  rob_summary, "Risk of Bias")
    if has_applic:
        _draw_summary_bar(ax_right, app_summary, "Applicability Concerns")
    else:
        ax_right.axis("off")

    _draw_legend(ax_legend)

    plt.tight_layout(pad=0.5)
    return _fig_to_b64(fig)

@login_required
@require_http_methods(['GET'])
def chart_info(request):
    project_id     = request.GET.get('project_id')
    quality_method = request.GET.get('quality_method', 'QUADAS2')
    if not project_id:
        return _json_err('缺少 project_id')
    try:
        chart = QAChart.objects.get(project_id=project_id, quality_method=quality_method)
    except QAChart.DoesNotExist:
        return _json_ok(None)   # 尚未生成
    return _json_ok({
        'id':             chart.id,
        'quality_method': chart.quality_method,
        'chart_types':    chart.chart_types,
        'ref_ids':        chart.ref_ids,
        'image_url':      chart.image_file.file.url if chart.image_file else None,
        'excel_url':      chart.excel_file.file.url if chart.excel_file else None,
        'generated_at':   chart.generated_at.isoformat() if chart.generated_at else None,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 导出
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
@login_required
@require_http_methods(['POST'])
def export_excel(request):
    """POST /api/qa/export/excel/ — 生成并返回 Excel 文件"""
    try:
        body = json.loads(request.body)
    except Exception:
        return _json_err('请求体 JSON 格式错误')

    project_id     = body.get('project_id')
    quality_method = body.get('quality_method', 'QUADAS2')
    include_unconfirmed = body.get('include_unconfirmed', False)

    if not project_id:
        return _json_err('缺少 project_id')
    project = _get_project(request, project_id)
    if not project:
        return _json_err('项目不存在', 404)

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return _json_err('服务端缺少 openpyxl 依赖')

    refs = QAReference.objects.filter(project=project, quality_method=quality_method).prefetch_related('signal_items__confirmed_by', 'domain_results')

    wb = openpyxl.Workbook()

    # ── Sheet 1: 评价明细 ──────────────────────────────────
    ws = wb.active
    ws.title = '评价明细'
    headers = [
        '项目名称', '文献标题', '第一作者', '年份', '质量评价方法', '评价模式',
        '评价领域', '结果类型', '信号问题', '中文释义',
        '模型1判断', '模型1理由', '模型2判断', '模型2理由',
        '一致性状态', '系统推荐', 'AI判断', '判断理由',
        '人工最终判断', '是否人工修改', '确认人', '确认时间', '证据位置',
        '是否已确认',
    ]
    ws.append(headers)

    COLOR_MAP = {'low': 'C8F7C5', 'high': 'FFCDD2', 'unclear': 'FFF9C4', 'pending': 'F5F5F5'}

    for ref in refs:
        for item in ref.signal_items.all():
            if not include_unconfirmed and not item.is_confirmed:
                row_color = 'F5F5F5'
            else:
                row_color = None

            row = [
                project.name,
                ref.title,
                ref.first_author,
                ref.year,
                ref.quality_method,
                ref.get_eval_mode_display() if ref.eval_mode else '',
                item.domain,
                item.result_type,
                item.signal_question,
                item.signal_description,
                item.model1_judgment,
                item.model1_reason,
                item.model2_judgment,
                item.model2_reason,
                item.consistency,
                item.system_recommendation,
                item.ai_judgment,
                item.ai_reason,
                item.human_judgment,
                '是' if item.is_modified else '否',
                item.confirmed_by.username if item.confirmed_by else '',
                item.confirmed_at.strftime('%Y-%m-%d %H:%M') if item.confirmed_at else '',
                item.ai_evidence_page,
                '是' if item.is_confirmed else '否（未确认）',
            ]
            ws.append(row)
            if row_color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(fill_type='solid', fgColor=row_color)

    # ── Sheet 2: 汇总统计 ──────────────────────────────────
    ws2 = wb.create_sheet('汇总统计')
    ws2.append(['文献标题', '第一作者', '年份', '患者选择_偏倚', '待评价试验_偏倚', '参考标准_偏倚', '流程与时间_偏倚',
                '患者选择_适用', '待评价试验_适用', '参考标准_适用', '整体审阅状态'])
    for ref in refs:
        dr_map = {dr.domain: dr for dr in ref.domain_results.all()}
        ps  = dr_map.get('patient_selection')
        it  = dr_map.get('index_test')
        rs  = dr_map.get('reference_standard')
        ft  = dr_map.get('flow_timing')
        ws2.append([
            ref.title, ref.first_author, ref.year,
            ps.bias_risk_result if ps else '',
            it.bias_risk_result if it else '',
            rs.bias_risk_result if rs else '',
            ft.bias_risk_result if ft else '',
            ps.applicability_result if ps else '',
            it.applicability_result if it else '',
            rs.applicability_result if rs else '',
            ref.review_status,
        ])

    # ── Sheet 3: 证据记录 ──────────────────────────────────
    ws3 = wb.create_sheet('证据记录')
    ws3.append(['文献标题', '信号问题', 'AI判断', '判断理由', '证据原文', '证据位置', '人工修改前', '人工最终判断', '是否修改'])
    for ref in refs:
        for item in ref.signal_items.filter(ai_evidence__gt=''):
            ws3.append([
                ref.title, item.signal_question,
                item.ai_judgment, item.ai_reason,
                item.ai_evidence, item.ai_evidence_page,
                item.original_ai_judgment, item.human_judgment,
                '是' if item.is_modified else '否',
            ])

    # ── Sheet 4: 多模型校验记录 ────────────────────────────
    ws4 = wb.create_sheet('多模型校验记录')
    ws4.append(['文献标题', '信号问题', '模型1 ID', '模型1判断', '模型1理由', '模型2 ID', '模型2判断', '模型2理由', '一致性', '系统推荐', '人工最终判断'])
    for ref in refs.filter(eval_mode__in=['multi', 'dual']):
        for item in ref.signal_items.exclude(consistency='single'):
            ws4.append([
                ref.title, item.signal_question,
                item.model1_id, item.model1_judgment, item.model1_reason,
                item.model2_id, item.model2_judgment, item.model2_reason,
                item.consistency, item.system_recommendation, item.human_judgment,
            ])

    # 返回文件流
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'qa_export_{project.name}_{quality_method}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'

    # ActivityLog
    from core.models import ActivityLog
    ActivityLog.objects.create(
        project=project,
        operation_type='qa_export_excel',
        operation_detail={
            'quality_method': quality_method,
            'include_unconfirmed': include_unconfirmed,
            'filename': filename,
        },
        created_by=request.user,
    )
    return resp


@login_required
@require_http_methods(['GET'])
def export_status(request):
    """GET /api/qa/export/status/?project_id="""
    project_id = request.GET.get('project_id')
    if not project_id:
        return _json_err('缺少 project_id')
    try:
        chart = QAChart.objects.filter(project_id=project_id).order_by('-created_at').first()
    except Exception:
        return _json_ok({'has_chart': False})

    if not chart:
        return _json_ok({'has_chart': False})

    return _json_ok({
        'has_chart':    True,
        'chart_id':     chart.id,
        'image_url':    chart.image_file.file.url if chart.image_file else None,
        'excel_url':    chart.excel_file.file.url if chart.excel_file else None,
        'generated_at': chart.generated_at.isoformat() if chart.generated_at else None,
    })
